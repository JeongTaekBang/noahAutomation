import pandas as pd
import openpyxl
import sys
import io

# UTF-8 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 파일 경로
file_path = r'C:\Users\Jeongtaek.Bang\OneDrive - Rotork plc\바탕 화면\업무\NOAH ACTUATION\NOAH_SO_PO_DN.xlsx'

def col_letter(n):
    """컬럼 번호를 엑셀 문자로 변환"""
    result = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

# Note 시트 확인
print('=== Note 시트 내용 ===')
try:
    df_note = pd.read_excel(file_path, sheet_name='Note')
    print(df_note.to_string())
except Exception as e:
    print(f'Error: {e}')

print('\n' + '='*50 + '\n')

# PO_국내 컬럼 구조
df_po = pd.read_excel(file_path, sheet_name='PO_국내', nrows=5)
df_so = pd.read_excel(file_path, sheet_name='SO_국내', nrows=5)

print('=== SO_국내 전체 컬럼 ===')
for i, c in enumerate(df_so.columns, 1):
    print(f'{col_letter(i)}: {c}')

print('\n=== PO_국내 전체 컬럼 ===')
for i, c in enumerate(df_po.columns, 1):
    print(f'{col_letter(i)}: {c}')

# openpyxl로 PO_국내의 수식 확인 (SO_국내 참조하는 컬럼들)
print('\n=== PO_국내 수식 분석 (2행 기준) ===')
wb = openpyxl.load_workbook(file_path, data_only=False)
ws_po = wb['PO_국내']

# 헤더 매핑
headers = {}
for col in range(1, 70):
    val = ws_po.cell(row=1, column=col).value
    if val:
        headers[col] = val

# SO_국내를 참조하는 수식 찾기
print('SO_국내를 참조하는 컬럼:')
for col, header in headers.items():
    cell = ws_po.cell(row=2, column=col)
    val = str(cell.value) if cell.value else ''
    if 'SO_' in val and '!' in val:  # 다른 시트 참조
        print(f'  {col_letter(col)}열 ({header}): {val}')

# 값만 있는 컬럼 (스냅샷)
print('\nPO_국내 주요 컬럼 상세 (2행 기준):')
key_cols = ['SO_ID', 'Customer name', 'Customer PO', 'Item name', 'Item qty', 'ICO Unit', 'Total ICO']
for col, header in headers.items():
    if header in key_cols:
        cell = ws_po.cell(row=2, column=col)
        val = cell.value

        # ArrayFormula 처리
        if hasattr(val, 'text'):
            print(f'  {col_letter(col)}열 ({header}): ArrayFormula = {val.text}')
        elif isinstance(val, str) and val.startswith('='):
            print(f'  {col_letter(col)}열 ({header}): 수식 = {val}')
        else:
            print(f'  {col_letter(col)}열 ({header}): 값 = {val}')

# 수식 참조 관계 전체 확인
print('\n=== PO_국내 참조 수식 전체 (2행) ===')
for col in range(1, 70):
    cell = ws_po.cell(row=2, column=col)
    val = cell.value
    header = headers.get(col, f'Col{col}')

    if hasattr(val, 'text'):  # ArrayFormula
        print(f'{col_letter(col)}열 ({header}): {val.text}')
    elif isinstance(val, str) and val.startswith('='):
        print(f'{col_letter(col)}열 ({header}): {val}')

wb.close()

# 행 번호 연동 확인
print('\n=== SO_국내 vs PO_국내 행 비교 (처음 5행) ===')
df_so = pd.read_excel(file_path, sheet_name='SO_국내', nrows=5)
df_po = pd.read_excel(file_path, sheet_name='PO_국내', nrows=5)

print('SO_국내:')
print(df_so[['SO_ID', 'Customer name', 'Item name', 'Item qty']].to_string())
print('\nPO_국내:')
print(df_po[['SO_ID', 'Customer name', 'Item name', 'Item qty']].to_string())
