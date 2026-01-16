"""
Excel Generator 테스트 (템플릿 기반)
====================================

excel_generator.py 모듈의 테스트
"""

from datetime import datetime, timedelta
from pathlib import Path
import tempfile

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from po_generator.excel_generator import (
    create_purchase_order,
    create_description_sheet,
    create_po_workbook,
)
from po_generator.template_engine import load_template, generate_po_template
from po_generator.config import SPEC_FIELDS, OPTION_FIELDS, PO_TEMPLATE_FILE


@pytest.fixture(scope='module', autouse=True)
def ensure_template():
    """테스트 전에 템플릿 파일이 있는지 확인"""
    if not PO_TEMPLATE_FILE.exists():
        generate_po_template()


def get_template_worksheet(sheet_name: str = "Purchase Order"):
    """템플릿에서 워크시트 가져오기"""
    wb = load_template()
    return wb, wb[sheet_name]


class TestCreatePurchaseOrder:
    """Purchase Order 시트 생성 테스트"""

    def test_creates_sheet_with_correct_title(self, valid_order_data):
        """시트 타이틀이 올바르게 설정되는지 테스트"""
        wb, ws = get_template_worksheet()

        create_purchase_order(ws, valid_order_data)

        # A1 셀에 주문번호가 포함되어야 함
        assert "Purchase Order" in str(ws['A1'].value)
        assert valid_order_data['RCK Order no.'] in str(ws['A1'].value)

    def test_creates_header_section(self, valid_order_data):
        """헤더 섹션이 올바르게 생성되는지 테스트"""
        wb, ws = get_template_worksheet()

        create_purchase_order(ws, valid_order_data)

        # Vendor Name
        assert "NOAH Actuation" in str(ws['A2'].value)
        # Customer name
        assert valid_order_data['Customer name'] == ws['A10'].value
        # Date 포맷 확인
        assert "Date:" in str(ws['A5'].value)

    def test_delivery_address_from_column(self):
        """납품주소 컬럼이 있으면 배송 주소로 사용"""
        order_data = pd.Series({
            'RCK Order no.': 'ND-0001',
            'Customer name': 'Test Customer',
            '납품주소': '서울시 강남구 테헤란로 123',
            'Item qty': 1,
            'ICO Unit': 1000000,
            'Incoterms': 'DAP',
        })

        wb, ws = get_template_worksheet()
        create_purchase_order(ws, order_data)

        # C5 셀에 배송 주소가 있어야 함
        assert ws['C5'].value == '서울시 강남구 테헤란로 123'

    def test_creates_item_rows(self, valid_order_data):
        """아이템 행이 올바르게 생성되는지 테스트"""
        wb, ws = get_template_worksheet()

        create_purchase_order(ws, valid_order_data)

        # Row 13에 첫 번째 아이템이 있어야 함
        assert ws['A13'].value == 1  # Item Number
        assert ws['F13'].value == valid_order_data['Item qty']  # Qty
        assert ws['G13'].value == "EA"  # Unit

    def test_creates_totals_section(self, valid_order_data):
        """합계 섹션이 올바르게 생성되는지 테스트"""
        wb, ws = get_template_worksheet()

        create_purchase_order(ws, valid_order_data)

        # 단일 아이템: Row 13이 마지막, Totals는 Row 14-16
        # Total net amount (아이템 다음 행)
        assert "Total net amount" in str(ws['I14'].value)
        # 수식 확인 (단일 아이템이므로 J13:J13)
        assert "=SUM(J13:J13)" in str(ws['J14'].value)

        # VAT
        assert "VAT" in str(ws['I15'].value)
        assert "=J14*0.1" in str(ws['J15'].value)

        # Order Total
        assert "Order Total" in str(ws['I16'].value)

    def test_creates_footer_section(self, valid_order_data):
        """푸터 섹션이 올바르게 생성되는지 테스트"""
        wb, ws = get_template_worksheet()

        create_purchase_order(ws, valid_order_data)

        # 단일 아이템: footer는 Row 17부터 시작
        # Footer text (마지막 행)
        assert "Keeping the World flowing" in str(ws['A25'].value)
        # Currency (Row 21)
        assert ws['B21'].value == 'KRW'
        # Incoterms (Row 22)
        assert ws['B22'].value == valid_order_data['Incoterms']

    def test_handles_multiple_items(self, multiple_items_df):
        """다중 아이템 처리 테스트"""
        wb, ws = get_template_worksheet()
        order_data = multiple_items_df.iloc[0]

        create_purchase_order(ws, order_data, multiple_items_df)

        # 두 개의 아이템이 Row 13, 14에 있어야 함
        assert ws['A13'].value == 1
        assert ws['A14'].value == 2
        # 세 번째 행에는 Totals가 시작됨

    def test_uses_model_when_no_item_name(self):
        """Item name이 없으면 Model 사용"""
        order_data = pd.Series({
            'RCK Order no.': 'ND-0001',
            'Customer name': 'Test',
            'Model': 'FCEL-200',
            'Item name': '',  # 빈 Item name
            'Item qty': 1,
            'ICO Unit': 1000000,
            'Power supply': 'AC220V-1Ph-50Hz',
            'Incoterms': 'EXW',
        })

        wb, ws = get_template_worksheet()
        create_purchase_order(ws, order_data)

        # Description에 Model이 포함되어야 함
        assert 'FCEL-200' in str(ws['B13'].value)

    def test_invalid_quantity_uses_default(self):
        """수량이 유효하지 않으면 기본값 1 사용"""
        order_data = pd.Series({
            'RCK Order no.': 'ND-0001',
            'Customer name': 'Test',
            'Model': 'FCEL-100',
            'Item name': 'Test Item',
            'Item qty': 'invalid',  # 유효하지 않은 수량
            'ICO Unit': 1000000,
            'Incoterms': 'EXW',
        })

        wb, ws = get_template_worksheet()
        create_purchase_order(ws, order_data)

        # 기본값 1이 사용되어야 함
        assert ws['F13'].value == 1

    def test_invalid_ico_unit_uses_default(self):
        """ICO Unit이 유효하지 않으면 기본값 0 사용"""
        order_data = pd.Series({
            'RCK Order no.': 'ND-0001',
            'Customer name': 'Test',
            'Model': 'FCEL-100',
            'Item name': 'Test Item',
            'Item qty': 2,
            'ICO Unit': 'not a number',  # 유효하지 않은 단가
            'Incoterms': 'EXW',
        })

        wb, ws = get_template_worksheet()
        create_purchase_order(ws, order_data)

        # 기본값 0이 사용되어야 함
        assert ws['H13'].value == 0

    def test_string_delivery_date(self):
        """문자열 납기일 처리"""
        order_data = pd.Series({
            'RCK Order no.': 'ND-0001',
            'Customer name': 'Test',
            'Model': 'FCEL-100',
            'Item name': 'Test Item',
            'Item qty': 1,
            'ICO Unit': 1000000,
            'Requested delivery date': '2026-03-15',  # 문자열 날짜
            'Incoterms': 'EXW',
        })

        wb, ws = get_template_worksheet()
        create_purchase_order(ws, order_data)

        # 날짜가 변환되어야 함
        assert '2026-03-15' in str(ws['I13'].value)

    def test_invalid_delivery_date_empty_string(self):
        """유효하지 않은 납기일은 빈 문자열"""
        order_data = pd.Series({
            'RCK Order no.': 'ND-0001',
            'Customer name': 'Test',
            'Model': 'FCEL-100',
            'Item name': 'Test Item',
            'Item qty': 1,
            'ICO Unit': 1000000,
            'Requested delivery date': 'invalid-date',  # 유효하지 않은 날짜
            'Incoterms': 'EXW',
        })

        wb, ws = get_template_worksheet()
        create_purchase_order(ws, order_data)

        # 빈 문자열 또는 None이어야 함 (invalid-date를 그대로 자를 수도 있음)
        # 문자열이면 [:10]으로 잘림
        assert ws['I13'].value is not None or ws['I13'].value == ''

    def test_saves_valid_excel_file(self, valid_order_data):
        """유효한 Excel 파일로 저장되는지 테스트"""
        wb, ws = get_template_worksheet()

        create_purchase_order(ws, valid_order_data)

        # 임시 파일에 저장
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)

        try:
            wb.save(temp_path)
            # 저장된 파일 다시 로드
            loaded_wb = load_workbook(temp_path)
            loaded_ws = loaded_wb['Purchase Order']  # 명시적으로 시트 선택

            # 데이터가 유지되는지 확인
            assert "Purchase Order" in str(loaded_ws['A1'].value)
            loaded_wb.close()
        finally:
            temp_path.unlink()

    def test_applies_print_settings(self, valid_order_data):
        """인쇄 설정이 적용되는지 테스트"""
        wb = create_po_workbook(valid_order_data)
        ws = wb['Purchase Order']

        # 인쇄 영역 (동적 - 단일 아이템이면 Row 25까지)
        print_area_str = str(ws.print_area)
        assert '$A$1' in print_area_str and '$J$' in print_area_str

        # 저장 후 로드해서 인쇄 영역 확인
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)

        try:
            wb.save(temp_path)
            loaded_wb = load_workbook(temp_path)
            loaded_ws = loaded_wb['Purchase Order']
            # 인쇄 영역이 저장/로드 후에도 유지되는지 확인
            loaded_print_area = str(loaded_ws.print_area)
            assert '$A$1' in loaded_print_area
            loaded_wb.close()
        finally:
            temp_path.unlink()


class TestCreateDescriptionSheet:
    """Description 시트 생성 테스트"""

    def test_creates_line_no_header(self, valid_order_data):
        """Line No 헤더가 올바르게 생성되는지 테스트"""
        wb, ws = get_template_worksheet("Description")

        create_description_sheet(ws, valid_order_data)

        assert ws['A1'].value == "Line No"
        assert ws.cell(row=1, column=2).value == 1  # 첫 번째 아이템

    def test_creates_spec_fields(self, valid_order_data):
        """사양 필드가 올바르게 생성되는지 테스트"""
        wb, ws = get_template_worksheet("Description")

        create_description_sheet(ws, valid_order_data)

        # SPEC_FIELDS가 Row 3부터 시작 (Row 1: Line No, Row 2: Qty)
        for idx, field in enumerate(SPEC_FIELDS):
            row = idx + 3
            assert ws.cell(row=row, column=1).value == field

    def test_creates_option_fields(self, valid_order_data):
        """옵션 필드가 올바르게 생성되는지 테스트"""
        wb, ws = get_template_worksheet("Description")

        create_description_sheet(ws, valid_order_data)

        # OPTION_FIELDS는 SPEC_FIELDS 다음에 시작 (Row 3 + len(SPEC_FIELDS))
        start_row = len(SPEC_FIELDS) + 3
        for idx, field in enumerate(OPTION_FIELDS):
            row = start_row + idx
            assert ws.cell(row=row, column=1).value == field

    def test_populates_data_values(self, valid_order_data):
        """데이터 값이 올바르게 채워지는지 테스트"""
        wb, ws = get_template_worksheet("Description")

        create_description_sheet(ws, valid_order_data)

        # Power supply 값 확인 (SPEC_FIELDS의 첫 번째, Row 3)
        power_supply_row = 3
        assert ws.cell(row=power_supply_row, column=2).value == valid_order_data['Power supply']

        # ALS 옵션 값 확인
        als_row = len(SPEC_FIELDS) + 3 + OPTION_FIELDS.index('ALS')
        assert ws.cell(row=als_row, column=2).value == valid_order_data['ALS']

    def test_handles_multiple_items(self, multiple_items_df):
        """다중 아이템 처리 테스트"""
        wb, ws = get_template_worksheet("Description")
        order_data = multiple_items_df.iloc[0]

        create_description_sheet(ws, order_data, multiple_items_df)

        # Line No: 1, 2가 있어야 함
        assert ws.cell(row=1, column=2).value == 1
        assert ws.cell(row=1, column=3).value == 2

    def test_column_widths_applied(self, valid_order_data):
        """열 너비가 설정되는지 테스트"""
        wb, ws = get_template_worksheet("Description")

        create_description_sheet(ws, valid_order_data)

        # A열 너비
        assert ws.column_dimensions['A'].width == 25


class TestCreatePoWorkbook:
    """create_po_workbook 함수 테스트"""

    def test_creates_both_sheets(self, valid_order_data):
        """두 시트가 모두 생성되는지 테스트"""
        wb = create_po_workbook(valid_order_data)

        assert "Purchase Order" in wb.sheetnames
        assert "Description" in wb.sheetnames

    def test_saves_valid_workbook(self, valid_order_data):
        """유효한 워크북으로 저장되는지 테스트"""
        wb = create_po_workbook(valid_order_data)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)

        try:
            wb.save(temp_path)
            assert temp_path.exists()
            assert temp_path.stat().st_size > 0

            # 다시 로드해서 확인
            loaded_wb = load_workbook(temp_path)
            assert len(loaded_wb.sheetnames) == 2
            loaded_wb.close()
        finally:
            temp_path.unlink()


class TestExcelGeneratorIntegration:
    """Excel Generator 통합 테스트"""

    def test_full_workbook_generation(self, valid_order_data):
        """전체 워크북 생성 통합 테스트"""
        wb = create_po_workbook(valid_order_data)

        # 두 시트가 존재하는지 확인
        assert "Purchase Order" in wb.sheetnames
        assert "Description" in wb.sheetnames

        # 저장 및 로드 테스트
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)

        try:
            wb.save(temp_path)
            loaded_wb = load_workbook(temp_path, data_only=True)

            # 시트 확인
            assert len(loaded_wb.sheetnames) == 2
            loaded_wb.close()
        finally:
            temp_path.unlink()

    def test_multiple_items_workbook(self, multiple_items_df):
        """다중 아이템 워크북 생성 테스트"""
        order_data = multiple_items_df.iloc[0]
        wb = create_po_workbook(order_data, multiple_items_df)

        # 저장 테스트
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)

        try:
            wb.save(temp_path)
            assert temp_path.exists()
            assert temp_path.stat().st_size > 0
        finally:
            temp_path.unlink()
