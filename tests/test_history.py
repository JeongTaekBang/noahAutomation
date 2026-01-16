"""
history 모듈 테스트

월별 폴더 구조: po_history/YYYY/M월/
"""

from pathlib import Path
from datetime import datetime
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import Workbook

from po_generator.history import (
    check_duplicate_order,
    save_to_history,
    get_history_count,
    get_all_history,
    clear_history,
    _sanitize_filename,
    _extract_data_from_po_file,
    get_current_month_info,
)


def create_mock_po_file(path: Path, order_no: str = "ND-0001", customer: str = "Test Customer") -> None:
    """테스트용 발주서 파일 생성 (실제 발주서 형식 - 동적 레이아웃)

    레이아웃 구조 (단일 아이템 기준):
    - Row 1-12: 헤더 영역 (고정)
    - Row 13: 아이템 행 (ITEM_START_ROW)
    - Row 14: Total Net Amount
    - Row 15: VAT
    - Row 16: Order Total
    - Row 17-22: 푸터 (Opportunity, Sector, Industry, Note, Currency, Incoterms)
    """
    wb = Workbook()

    # Purchase Order 시트
    ws_po = wb.active
    ws_po.title = "Purchase Order"
    ws_po['A1'] = f"Purchase Order - {order_no}"
    ws_po['A5'] = "Date:  11/JAN/2026"
    ws_po['A10'] = customer

    # 아이템 정보 (Row 13)
    ws_po['B13'] = "FCEL-100, AC220V, ALS"
    ws_po['F13'] = 2
    ws_po['H13'] = 1500000
    ws_po['I13'] = "2026-02-15"

    # 합계 섹션 (동적 위치: 아이템 1개일 때 Row 14-16)
    ws_po['J14'] = 3000000  # Total Net Amount
    ws_po['J15'] = 300000   # VAT
    ws_po['J16'] = 3300000  # Order Total

    # 푸터 섹션 (동적 위치: 아이템 1개일 때 Row 17-22)
    ws_po['D17'] = "Project A"      # Opportunity
    ws_po['D18'] = "Oil & Gas"      # Sector
    ws_po['D19'] = "IND001"         # Industry code
    ws_po['C20'] = "Note. Test"     # Remark
    ws_po['B21'] = "KRW"            # Currency
    ws_po['B22'] = "DAP"            # Incoterms

    # Description 시트
    ws_desc = wb.create_sheet("Description")
    ws_desc['A1'] = "Line No"
    ws_desc['B1'] = 1
    ws_desc['A2'] = "Power supply"
    ws_desc['B2'] = "AC220V-1Ph-50Hz"
    ws_desc['A3'] = "Motor(kW)"
    ws_desc['B3'] = "0.18"
    ws_desc['A4'] = "Model"
    ws_desc['B4'] = "FCEL-100"
    ws_desc['A5'] = "ALS"
    ws_desc['B5'] = "Y"

    wb.save(path)


class TestSanitizeFilename:
    """파일명 sanitize 테스트"""

    def test_special_characters(self):
        """특수문자 제거"""
        assert _sanitize_filename('A/B:C*D') == 'A_B_C_D'

    def test_spaces(self):
        """공백 처리"""
        assert _sanitize_filename('Test  Name') == 'Test_Name'

    def test_korean(self):
        """한글 유지"""
        assert _sanitize_filename('고객사') == '고객사'


class TestExtractDataFromPoFile:
    """발주서 데이터 추출 테스트"""

    def test_extract_purchase_order_data(self, tmp_path: Path):
        """Purchase Order 시트 데이터 추출"""
        po_file = tmp_path / 'test_po.xlsx'
        create_mock_po_file(po_file, 'ND-0001', 'Test Customer')

        record = _extract_data_from_po_file(po_file)

        assert record.get('Customer name') == 'Test Customer'
        assert record.get('Currency') == 'KRW'
        assert record.get('Incoterms') == 'DAP'
        assert record.get('Item count') == 1

    def test_extract_description_data(self, tmp_path: Path):
        """Description 시트 데이터 추출"""
        po_file = tmp_path / 'test_po.xlsx'
        create_mock_po_file(po_file, 'ND-0001', 'Test Customer')

        record = _extract_data_from_po_file(po_file)

        assert record.get('Power supply') == 'AC220V-1Ph-50Hz'
        assert record.get('Motor(kW)') == '0.18'
        assert record.get('Model') == 'FCEL-100'
        assert record.get('ALS') == 'Y'

    def test_corrupted_file_returns_empty(self, tmp_path: Path):
        """손상된 파일은 빈 딕셔너리 반환"""
        po_file = tmp_path / 'corrupted.xlsx'
        po_file.write_text('not a valid xlsx file')

        record = _extract_data_from_po_file(po_file)
        assert record == {}

    def test_file_without_sheets_returns_partial(self, tmp_path: Path):
        """시트 없는 파일은 부분 데이터 반환"""
        po_file = tmp_path / 'empty.xlsx'
        # 빈 워크북 생성 (시트 없음은 불가하므로 다른 이름의 시트만 있는 경우)
        wb = Workbook()
        wb.active.title = "OtherSheet"
        wb.save(po_file)

        record = _extract_data_from_po_file(po_file)
        # Purchase Order, Description 시트가 없으므로 빈 딕셔너리
        assert record == {}


class TestCheckDuplicateOrder:
    """중복 발주 체크 테스트 (현재 월 폴더만)"""

    def test_no_history(self, tmp_path: Path):
        """이력이 없는 경우"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            # check_all_months=False로 현재 월만 검색
            result = check_duplicate_order('ND-0001', check_all_months=False)
            assert result is None

    def test_order_in_folder(self, tmp_path: Path):
        """현재 월 폴더에 있는 주문번호"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        # 폴더에 DB 형식 이력 파일 생성
        test_file = month_dir / '20260101_ND-0001_TestCustomer.xlsx'
        df = pd.DataFrame([{
            '생성일시': '2026-01-01 10:00:00',
            'RCK Order no.': 'ND-0001',
            'Customer name': 'Test Customer',
        }])
        df.to_excel(test_file, index=False)

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            # check_all_months=False로 현재 월만 검색
            result = check_duplicate_order('ND-0001', check_all_months=False)
            assert result is not None
            assert result['생성일시'] == '2026-01-01 10:00:00'

    def test_order_not_in_history(self, tmp_path: Path):
        """이력에 없는 주문번호"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        # 다른 주문번호로 이력 생성
        test_file = month_dir / '20260101_ND-0001_TestCustomer.xlsx'
        df = pd.DataFrame([{'RCK Order no.': 'ND-0001'}])
        df.to_excel(test_file, index=False)

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            result = check_duplicate_order('ND-9999', check_all_months=False)
            assert result is None

    def test_check_all_months(self, tmp_path: Path):
        """전체 월 검색 테스트"""
        history_dir = tmp_path / 'po_history'
        # 이전 월 폴더에 이력 생성
        prev_month_dir = history_dir / '2025' / '12월'
        prev_month_dir.mkdir(parents=True)

        test_file = prev_month_dir / '20251215_ND-0001_TestCustomer.xlsx'
        df = pd.DataFrame([{
            '생성일시': '2025-12-15 10:00:00',
            'RCK Order no.': 'ND-0001',
        }])
        df.to_excel(test_file, index=False)

        with patch('po_generator.history.HISTORY_DIR', history_dir):
            # 전체 검색 시 이전 월에서도 찾아야 함
            result = check_duplicate_order('ND-0001', check_all_months=True)
            assert result is not None
            assert '2025-12-15' in result['생성일시']

    def test_corrupted_history_file_uses_mtime(self, tmp_path: Path):
        """손상된 이력 파일은 파일 수정 시간 사용"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        # 손상된 파일 생성
        test_file = month_dir / '20260101_ND-0001_TestCustomer.xlsx'
        test_file.write_text('corrupted content')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            result = check_duplicate_order('ND-0001', check_all_months=False)
            assert result is not None
            assert '생성일시' in result
            assert '생성파일' in result

    def test_skips_non_directory_in_search(self, tmp_path: Path):
        """검색 중 디렉토리가 아닌 항목은 건너뜀"""
        history_dir = tmp_path / 'po_history'
        history_dir.mkdir(parents=True)

        # "1월"이라는 파일 생성 (디렉토리 아님)
        fake_month = history_dir / '2026' / '1월'
        fake_month.parent.mkdir(parents=True)
        fake_month.write_text('this is a file, not a directory')

        with patch('po_generator.history.HISTORY_DIR', history_dir):
            # 오류 없이 None 반환
            result = check_duplicate_order('ND-0001', check_all_months=True)
            assert result is None


class TestSaveToHistory:
    """이력 저장 테스트 (월별 폴더)"""

    def test_save_creates_folder(self, tmp_path: Path):
        """월별 폴더가 없으면 생성"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        output_file = tmp_path / 'output.xlsx'

        create_mock_po_file(output_file, 'ND-TEST-001', 'Test Customer')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            result = save_to_history(output_file, 'ND-TEST-001', 'Test Customer')
            assert result is True
            assert month_dir.exists()

    def test_save_db_format(self, tmp_path: Path):
        """DB 형식(한 행)으로 저장 확인"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        output_file = tmp_path / 'output.xlsx'

        create_mock_po_file(output_file, 'ND-TEST-001', 'Test Customer')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            save_to_history(output_file, 'ND-TEST-001', 'Test Customer')

            # 생성된 파일 확인
            files = list(month_dir.glob('*.xlsx'))
            assert len(files) == 1

            df = pd.read_excel(files[0])
            assert len(df) == 1  # DB 형식: 한 행

            # 메타 정보 확인
            assert '생성일시' in df.columns
            assert 'RCK Order no.' in df.columns
            assert 'Customer name' in df.columns
            assert '원본파일' in df.columns

            # 추출된 데이터 확인
            assert 'Power supply' in df.columns
            assert 'Model' in df.columns
            assert df.iloc[0]['Model'] == 'FCEL-100'

    def test_filename_format(self, tmp_path: Path):
        """파일명 형식 확인"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        output_file = tmp_path / 'output.xlsx'

        create_mock_po_file(output_file, 'ND-TEST-001', 'Test Customer')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            save_to_history(output_file, 'ND-TEST-001', 'Test Customer')

            files = list(month_dir.glob('*.xlsx'))
            filename = files[0].stem

            today = datetime.now().strftime("%Y%m%d")
            assert filename.startswith(today)
            assert 'ND-TEST-001' in filename
            assert 'Test_Customer' in filename

    def test_duplicate_filename_handling(self, tmp_path: Path):
        """동일 날짜에 같은 주문 시 번호 추가"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)
        output_file = tmp_path / 'output.xlsx'

        create_mock_po_file(output_file, 'ND-TEST-001', 'Test Customer')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            save_to_history(output_file, 'ND-TEST-001', 'Test Customer')
            save_to_history(output_file, 'ND-TEST-001', 'Test Customer')

            files = list(month_dir.glob('*.xlsx'))
            assert len(files) == 2

    def test_nonexistent_file(self, tmp_path: Path):
        """존재하지 않는 파일"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        output_file = tmp_path / 'nonexistent.xlsx'

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            result = save_to_history(output_file, 'ND-TEST-001', 'Test Customer')
            assert result is False

    def test_save_failure_returns_false(self, tmp_path: Path):
        """저장 실패 시 False 반환"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        output_file = tmp_path / 'output.xlsx'

        create_mock_po_file(output_file, 'ND-TEST-001', 'Test Customer')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            with patch('pandas.DataFrame.to_excel', side_effect=PermissionError("Access denied")):
                result = save_to_history(output_file, 'ND-TEST-001', 'Test Customer')
                assert result is False


class TestGetHistoryCount:
    """이력 건수 조회 테스트 (현재 월 폴더만)"""

    def test_no_history(self, tmp_path: Path):
        """이력이 없는 경우"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            count = get_history_count()
            assert count == 0

    def test_folder_count(self, tmp_path: Path):
        """현재 월 폴더 건수"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        # 3개 파일 생성
        for i in range(3):
            f = month_dir / f'20260101_ND-000{i}_Customer.xlsx'
            pd.DataFrame([{'test': i}]).to_excel(f, index=False)

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            count = get_history_count()
            assert count == 3


class TestGetAllHistory:
    """현재 월 이력 조회 테스트"""

    def test_empty_history(self, tmp_path: Path):
        """이력이 없는 경우"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            df = get_all_history()
            assert df.empty

    def test_merge_db_files(self, tmp_path: Path):
        """DB 형식 파일들 합치기"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        # 2개 DB 형식 파일 생성
        for i in range(2):
            f = month_dir / f'20260101_ND-000{i}_Customer.xlsx'
            pd.DataFrame([{
                '생성일시': f'2026-01-0{i+1} 10:00:00',
                'RCK Order no.': f'ND-000{i}',
                'Model': f'FCEL-{i}00',
            }]).to_excel(f, index=False)

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            df = get_all_history()
            assert len(df) == 2
            assert 'Model' in df.columns

    def test_skips_corrupted_files(self, tmp_path: Path):
        """손상된 파일은 건너뜀"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        # 정상 파일
        valid_file = month_dir / '20260101_ND-0001_Customer.xlsx'
        pd.DataFrame([{'RCK Order no.': 'ND-0001'}]).to_excel(valid_file, index=False)

        # 손상된 파일
        corrupted_file = month_dir / '20260102_ND-0002_Customer.xlsx'
        corrupted_file.write_text('not a valid xlsx')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            df = get_all_history()
            # 정상 파일만 읽힘
            assert len(df) == 1
            assert df.iloc[0]['RCK Order no.'] == 'ND-0001'


class TestClearHistory:
    """이력 초기화 테스트 (현재 월 폴더만)"""

    def test_clear_folder(self, tmp_path: Path):
        """현재 월 폴더 내 파일 삭제"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        (month_dir / 'test1.xlsx').write_text('dummy')
        (month_dir / 'test2.xlsx').write_text('dummy')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            result = clear_history()
            assert result is True
            assert len(list(month_dir.glob('*.xlsx'))) == 0

    def test_clear_nonexistent(self, tmp_path: Path):
        """없는 이력 삭제 시도"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            result = clear_history()
            assert result is True

    def test_clear_permission_error(self, tmp_path: Path):
        """삭제 권한 없을 때 False 반환"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        (month_dir / 'test.xlsx').write_text('dummy')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            with patch.object(Path, 'glob', side_effect=PermissionError("Access denied")):
                result = clear_history()
                assert result is False


class TestGetCurrentMonthInfo:
    """현재 월 정보 테스트"""

    def test_returns_tuple(self):
        """튜플 반환 확인"""
        month_str, month_dir = get_current_month_info()
        assert isinstance(month_str, str)
        assert isinstance(month_dir, Path)

    def test_format(self):
        """형식 확인"""
        month_str, month_dir = get_current_month_info()
        now = datetime.now()
        assert f"{now.year}년" in month_str
        assert f"{now.month}월" in month_str
        assert str(now.year) in str(month_dir)
        assert f"{now.month}월" in str(month_dir)
