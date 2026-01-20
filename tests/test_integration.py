"""
통합 테스트 모듈
================

여러 모듈에 걸친 통합 테스트를 제공합니다:
- 행 삭제/삽입 동작 검증
- 이력 파일명 시퀀스 검증
- find_item_start_row 일관성 검증
"""

from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import Workbook

from po_generator.history import (
    save_to_history,
    _get_history_filename,
)
from po_generator.excel_helpers import (
    find_item_start_row_openpyxl,
    find_item_start_row,
    PO_HEADER_LABELS,
    TS_HEADER_LABELS,
    DEFAULT_HEADER_LABELS,
)


def create_mock_po_file_with_items(path: Path, order_no: str, customer: str, item_count: int = 1) -> None:
    """테스트용 발주서 파일 생성 (가변 아이템 수)

    Args:
        path: 파일 경로
        order_no: 주문 번호
        customer: 고객명
        item_count: 아이템 수
    """
    wb = Workbook()

    # Purchase Order 시트
    ws_po = wb.active
    ws_po.title = "Purchase Order"
    ws_po['A1'] = f"Purchase Order - {order_no}"
    ws_po['A5'] = "Date:  11/JAN/2026"
    ws_po['A10'] = customer

    # 아이템 헤더 (Row 12)
    ws_po['A12'] = "No."
    ws_po['B12'] = "Item Number"

    # 아이템 정보 (Row 13부터)
    for i in range(item_count):
        row = 13 + i
        ws_po[f'A{row}'] = i + 1
        ws_po[f'B{row}'] = f"FCEL-{100 + i}, AC220V, ALS"
        ws_po[f'F{row}'] = 2
        ws_po[f'H{row}'] = 1500000
        ws_po[f'I{row}'] = "2026-02-15"

    # 합계 섹션 (아이템 마지막 행 + 1)
    total_row = 13 + item_count
    ws_po[f'J{total_row}'] = 3000000 * item_count  # Total Net Amount
    ws_po[f'J{total_row + 1}'] = 300000 * item_count  # VAT
    ws_po[f'J{total_row + 2}'] = 3300000 * item_count  # Order Total

    # 푸터 섹션
    footer_row = total_row + 3
    ws_po[f'D{footer_row}'] = "Project A"
    ws_po[f'D{footer_row + 1}'] = "Oil & Gas"
    ws_po[f'D{footer_row + 2}'] = "IND001"
    ws_po[f'C{footer_row + 3}'] = "Note. Test"
    ws_po[f'B{footer_row + 4}'] = "KRW"
    ws_po[f'B{footer_row + 5}'] = "DAP"

    # Description 시트
    ws_desc = wb.create_sheet("Description")
    ws_desc['A1'] = "Line No"
    ws_desc['B1'] = 1
    ws_desc['A2'] = "Power supply"
    ws_desc['B2'] = "AC220V-1Ph-50Hz"
    ws_desc['A3'] = "Model"
    ws_desc['B3'] = "FCEL-100"

    wb.save(path)


class TestHistoryFilenameSequence:
    """이력 파일명 중복 시 순번 추가 테스트"""

    def test_first_save_no_counter(self, tmp_path: Path):
        """첫 번째 저장 시 카운터 없음"""
        filename = _get_history_filename('ND-0001', 'Test Customer')
        today = datetime.now().strftime("%Y%m%d")

        assert filename.startswith(today)
        assert 'ND-0001' in filename
        assert 'Test_Customer' in filename
        # 첫 번째 저장이므로 _1, _2 없음
        assert '_1' not in filename
        assert '_2' not in filename

    def test_duplicate_adds_counter(self, tmp_path: Path):
        """동일 날짜에 같은 주문 저장 시 _1, _2 추가"""
        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        output_file = tmp_path / 'output.xlsx'
        create_mock_po_file_with_items(output_file, 'ND-0001', 'Test Customer')

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            # 첫 번째 저장
            result1 = save_to_history(output_file, 'ND-0001', 'Test Customer')
            assert result1 is True

            # 두 번째 저장
            result2 = save_to_history(output_file, 'ND-0001', 'Test Customer')
            assert result2 is True

            # 세 번째 저장
            result3 = save_to_history(output_file, 'ND-0001', 'Test Customer')
            assert result3 is True

            # 파일 목록 확인
            files = sorted(month_dir.glob('*.xlsx'))
            assert len(files) == 3

            # 파일명 확인
            filenames = [f.stem for f in files]
            today = datetime.now().strftime("%Y%m%d")

            # 첫 번째: 카운터 없음
            assert any(f'{today}_ND-0001_Test_Customer' == name for name in filenames)
            # 두 번째: _1
            assert any(f'{today}_ND-0001_Test_Customer_1' == name for name in filenames)
            # 세 번째: _2
            assert any(f'{today}_ND-0001_Test_Customer_2' == name for name in filenames)


class TestFindItemStartRowConsistency:
    """find_item_start_row openpyxl/xlwings 일관성 테스트"""

    def test_po_header_labels_openpyxl(self, tmp_path: Path):
        """PO 헤더 라벨로 openpyxl에서 아이템 시작 행 찾기"""
        test_file = tmp_path / 'test.xlsx'
        create_mock_po_file_with_items(test_file, 'ND-0001', 'Test', item_count=1)

        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        ws = wb['Purchase Order']

        # Row 12에 "No." 헤더가 있으므로 Row 13 반환
        result = find_item_start_row_openpyxl(ws, PO_HEADER_LABELS)
        assert result == 13

        wb.close()

    def test_fallback_when_header_not_found(self, tmp_path: Path):
        """헤더를 찾지 못할 때 fallback 값 반환"""
        test_file = tmp_path / 'test.xlsx'
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Some random content"
        ws['B2'] = "No header here"
        wb.save(test_file)

        wb_read = __import__('openpyxl').load_workbook(test_file)
        ws_read = wb_read.active

        # 헤더 없으므로 fallback 값 반환
        result = find_item_start_row_openpyxl(ws_read, PO_HEADER_LABELS, fallback_row=99)
        assert result == 99

        wb_read.close()

    def test_ts_header_labels(self, tmp_path: Path):
        """거래명세표 헤더 라벨로 아이템 시작 행 찾기"""
        test_file = tmp_path / 'test.xlsx'
        wb = Workbook()
        ws = wb.active

        # 거래명세표 형식 시뮬레이션
        ws['A10'] = "월/일"
        ws['B10'] = "품명"
        wb.save(test_file)

        wb_read = __import__('openpyxl').load_workbook(test_file)
        ws_read = wb_read.active

        result = find_item_start_row_openpyxl(ws_read, TS_HEADER_LABELS)
        assert result == 11  # 헤더 Row 10 + 1

        wb_read.close()

    def test_auto_detect_openpyxl(self, tmp_path: Path):
        """find_item_start_row가 openpyxl을 자동 감지"""
        test_file = tmp_path / 'test.xlsx'
        create_mock_po_file_with_items(test_file, 'ND-0001', 'Test', item_count=1)

        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        ws = wb['Purchase Order']

        # 자동 감지 함수 사용
        result = find_item_start_row(ws, PO_HEADER_LABELS)
        assert result == 13

        wb.close()


class TestExcelHelperLabels:
    """excel_helpers 라벨 프리셋 테스트"""

    def test_po_header_labels_exist(self):
        """PO 헤더 라벨 프리셋 존재"""
        assert 'No.' in PO_HEADER_LABELS
        assert 'Item Number' in PO_HEADER_LABELS

    def test_ts_header_labels_exist(self):
        """TS 헤더 라벨 프리셋 존재"""
        assert '월/일' in TS_HEADER_LABELS
        assert '품명' in TS_HEADER_LABELS

    def test_default_header_labels_comprehensive(self):
        """기본 헤더 라벨 프리셋이 포괄적"""
        assert 'No.' in DEFAULT_HEADER_LABELS
        assert 'Item Number' in DEFAULT_HEADER_LABELS
        assert '품명' in DEFAULT_HEADER_LABELS


class TestRowDeletionBehavior:
    """행 삭제 동작 테스트 (문서화용)

    이 테스트는 xlUp 삭제 동작을 문서화합니다:
    - 같은 위치에서 반복 삭제 시 아래 행이 올라옴
    - 예: Row 15, 16, 17 삭제 시 → Row 15에서 3번 삭제
    """

    def test_xlup_behavior_documented(self):
        """xlUp 삭제 동작 문서화

        시나리오: 템플릿에 3개 아이템 행이 있고 1개만 필요한 경우
        - 아이템 시작 행: 13
        - 필요 아이템: 1개
        - 삭제할 행: Row 14, 15 (13 + 1 = 14에서 시작)

        xlUp 동작:
        1. Row 14 삭제 → Row 15, 16이 14, 15로 올라옴
        2. Row 14 삭제 → (원래 Row 16이) Row 15로 올라옴

        따라서 delete_row = item_start_row + num_items (고정 위치)에서
        반복 삭제하면 연속된 행들이 삭제됨
        """
        # 이 테스트는 xlUp 동작의 문서화 목적
        # 실제 xlwings 테스트는 Excel COM 인터페이스가 필요하므로
        # 여기서는 동작 원리만 검증

        item_start_row = 13
        num_items = 1
        template_item_count = 3
        rows_to_delete = template_item_count - num_items

        # 삭제할 행 계산
        delete_row = item_start_row + num_items
        assert delete_row == 14
        assert rows_to_delete == 2

        # 2번 반복하면 원래 Row 14, 15가 삭제됨
        # (Row 14 삭제 → Row 14에 원래 Row 15 내용
        #  Row 14 삭제 → Row 14에 원래 Row 16 내용)


class TestValidOrderDataIntegration:
    """유효한 주문 데이터 통합 테스트"""

    def test_save_and_retrieve_history(self, tmp_path: Path):
        """이력 저장 후 조회"""
        from po_generator.history import get_all_history

        month_dir = tmp_path / 'po_history' / '2026' / '1월'
        month_dir.mkdir(parents=True)

        output_file = tmp_path / 'output.xlsx'
        create_mock_po_file_with_items(output_file, 'ND-0001', 'Test Customer', item_count=2)

        with patch('po_generator.history._get_current_month_dir', return_value=month_dir):
            # 저장
            result = save_to_history(output_file, 'ND-0001', 'Test Customer')
            assert result is True

            # 조회
            df = get_all_history()
            assert len(df) == 1
            assert df.iloc[0]['RCK Order no.'] == 'ND-0001'
            assert df.iloc[0]['Customer name'] == 'Test Customer'
            # 아이템 수 확인 (2개)
            assert df.iloc[0]['Item count'] == 2
