#!/usr/bin/env python
"""
SO 매출대사 (Sales Reconciliation)
===================================

AX ERP 매출 금액과 NOAH DN(매출) 금액을
AX Project 기준으로 비교하여 금액 일치 여부를 확인합니다.

사용법:
    python reconcile_so.py P03           # 3월 대사
    python reconcile_so.py P03 -v        # 상세 로그
"""

from __future__ import annotations

import argparse
import logging
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

from po_generator.config import (
    NOAH_SO_PO_DN_FILE, BASE_DIR,
    DN_DOMESTIC_SHEET, DN_EXPORT_SHEET,
)
from po_generator.logging_config import setup_logging
from po_generator.recon_paths import resolve_period_dir

logger = logging.getLogger(__name__)

RECON_DIR = BASE_DIR / "so_reconciliation"

# DN 시트에서 사용할 컬럼 (국내/해외 공통 + 해외 전용)
# 국내: AX Project 컬럼명 = 'AX Project no'
# 해외: AX Project 컬럼명 = 'AX Project number'
AX_PROJECT_COL = 'AX Project number'  # 통합 후 사용할 컬럼명

DN_COMMON_COLS = [
    'DN_ID', 'SO_ID', 'Customer name', 'Item', 'Line item',
    'Qty', 'Currency', 'Unit Price', 'Total Sales', '출고일',
]

DN_EXPORT_EXTRA_COLS = ['Total Sales KRW', '선적일']

FX_SHEET = 'FX'
# 환율차이 판정 임계값 (반올림 오차 허용)
FX_DIFF_THRESHOLD = 100


def find_ax_sales_file(period_code: str) -> Path | None:
    """period 디렉터리(플랫 또는 연도 중첩)에서 AX_Sales 파일 찾기"""
    period_dir = resolve_period_dir(RECON_DIR, period_code)
    if period_dir is None:
        return None
    for f in period_dir.iterdir():
        if f.suffix == '.xlsx' and not f.name.startswith('~'):
            if 'AX_SALES' in f.name.upper():
                return f
    return None


def load_ax_sales(file_path: Path) -> pd.DataFrame:
    """AX Sales 파일 로드 (Project, Customer, AX 금액)"""
    xf = pd.ExcelFile(file_path)

    # 시트명: Sheet1 우선, Sales fallback, 그 외 첫 번째 시트
    if 'Sheet1' in xf.sheet_names:
        sheet = 'Sheet1'
    elif 'Sales' in xf.sheet_names:
        sheet = 'Sales'
    else:
        sheet = xf.sheet_names[0]

    df = pd.read_excel(xf, sheet_name=sheet)

    # 필수 컬럼 확인
    for col in ['Project', 'AX']:
        if col not in df.columns:
            raise ValueError(f"AX Sales 파일에 '{col}' 컬럼이 없습니다: {list(df.columns)}")

    if 'Customer' not in df.columns:
        df['Customer'] = ''

    df = df[df['Project'].notna()].copy()
    df['Project'] = df['Project'].astype(str).str.strip()
    df['AX'] = pd.to_numeric(df['AX'], errors='coerce').fillna(0)

    # NOAH_SO 컬럼은 무시 (자동 계산)
    df = df[['Project', 'Customer', 'AX']].copy()

    # Project 중복 검출 및 집계 (NOAH 측 aggregate_noah_dn과 동일 패턴)
    dup_mask = df.duplicated(subset=['Project'], keep=False)
    if dup_mask.any():
        dup_counts = df[dup_mask].groupby('Project').size()
        logger.warning(
            "AX Sales 중복 Project %d건 — 합산 처리: %s",
            len(dup_counts), dict(dup_counts),
        )
        df = df.groupby('Project', as_index=False).agg({
            'AX': 'sum', 'Customer': 'first',
        })

    logger.debug("AX Sales 로드: %d건 (%s)", len(df), file_path.name)
    return df


def load_noah_dn(year_month: str) -> pd.DataFrame:
    """NOAH_SO_PO_DN.xlsx에서 DN_국내/DN_해외 읽기 (매출 데이터)

    Args:
        year_month: 출고일 필터 기준 (예: '2026-03')
    """
    xf = pd.ExcelFile(NOAH_SO_PO_DN_FILE)

    # DN_국내
    df_dom = pd.read_excel(xf, DN_DOMESTIC_SHEET)
    # 국내는 AX Project 컬럼명이 'AX Project no'
    if 'AX Project no' in df_dom.columns:
        df_dom = df_dom.rename(columns={'AX Project no': AX_PROJECT_COL})
    dom_cols = [c for c in [AX_PROJECT_COL] + DN_COMMON_COLS if c in df_dom.columns]
    df_dom = df_dom[dom_cols].copy()
    df_dom['구분'] = '국내'
    # 국내는 Total Sales = KRW
    df_dom['Total Sales KRW'] = df_dom['Total Sales']

    # DN_해외
    df_exp = pd.read_excel(xf, DN_EXPORT_SHEET)
    exp_cols = [c for c in [AX_PROJECT_COL] + DN_COMMON_COLS + DN_EXPORT_EXTRA_COLS
                if c in df_exp.columns]
    df_exp = df_exp[exp_cols].copy()
    df_exp['구분'] = '해외'

    # 매출 인식일: 국내=출고일, 해외=선적일
    df_dom['매출일'] = pd.to_datetime(df_dom['출고일'], errors='coerce')
    if '선적일' in df_exp.columns:
        df_exp['매출일'] = pd.to_datetime(df_exp['선적일'], errors='coerce')
    else:
        df_exp['매출일'] = pd.to_datetime(df_exp['출고일'], errors='coerce')

    # 통합
    df_all = pd.concat([df_dom, df_exp], ignore_index=True)

    # AX Project number 있는 건만
    has_proj = df_all[AX_PROJECT_COL].notna() & (df_all[AX_PROJECT_COL] != '')
    df_all = df_all[has_proj].copy()
    df_all[AX_PROJECT_COL] = df_all[AX_PROJECT_COL].astype(str).str.strip()

    # 매출일 기준 해당 월 필터
    df_all['매출월'] = df_all['매출일'].dt.to_period('M').astype(str)
    before = len(df_all)
    df_all = df_all[df_all['매출월'] == year_month].copy()
    logger.debug("NOAH DN 로드: %d건 → 매출일 %s 필터 → %d건",
                 before, year_month, len(df_all))

    return df_all


def load_fx_rates() -> pd.DataFrame:
    """FX 시트에서 환율 테이블 로드

    Returns:
        DataFrame with index=Currency (USD/EUR/GBP), columns=period (2026-01, ...)
    """
    df = pd.read_excel(NOAH_SO_PO_DN_FILE, sheet_name=FX_SHEET)
    df = df.set_index('FX')
    logger.debug("FX 환율 로드: %s", list(df.columns))
    return df


def period_code_to_col(period_code: str, fx_cols: list[str]) -> str | None:
    """P03 → FX 시트의 매칭 컬럼명 (예: '2026-03') 반환 — 다년 시 최신 연도"""
    month = period_code.replace('P', '').zfill(2)  # P03 → '03'
    matches = [col for col in fx_cols if str(col).endswith(f'-{month}')]
    if not matches:
        return None
    if len(matches) > 1:
        logger.warning(
            "FX 시트에 %s월 컬럼 %d개 — 최신 사용: %s", month, len(matches), matches,
        )
    return max(matches, key=str)  # YYYY-MM 포맷은 사전순 = 연도순


def aggregate_noah_dn(df_dn: pd.DataFrame) -> pd.DataFrame:
    """DN 데이터를 AX Project number 기준으로 집계"""
    agg = (df_dn.groupby(AX_PROJECT_COL)
           .agg(
               Customer=('Customer name', 'first'),
               구분=('구분', 'first'),
               Currency=('Currency', 'first'),
               NOAH_금액_KRW=('Total Sales KRW', 'sum'),
               외화금액=('Total Sales', 'sum'),
               건수=(AX_PROJECT_COL, 'size'),
           )
           .reset_index())

    # 환율: 해외만 계산 (= 적용 환율 역산)
    agg['환율'] = np.where(
        (agg['Currency'] != 'KRW') & (agg['외화금액'] > 0),
        agg['NOAH_금액_KRW'] / agg['외화금액'],
        np.nan,
    )

    # 국내: 외화금액/환율은 의미 없으므로 NaN
    is_krw = agg['Currency'] == 'KRW'
    agg.loc[is_krw, '외화금액'] = np.nan
    agg.loc[is_krw, '환율'] = np.nan

    return agg


def build_reconciliation(
    ax_sales: pd.DataFrame,
    noah_agg: pd.DataFrame,
    fx_rates: pd.DataFrame | None,
    recon_period_col: str | None,
) -> pd.DataFrame:
    """AX Sales 기준으로 NOAH DN 대사 결과 생성

    Args:
        ax_sales: AX Sales DataFrame
        noah_agg: NOAH DN 집계 DataFrame
        fx_rates: FX 환율 테이블 (index=Currency, columns=period)
        recon_period_col: 대사 월에 해당하는 FX 컬럼명 (예: '2026-03')

    Returns:
        summary DataFrame
    """
    merged = ax_sales.merge(
        noah_agg,
        left_on='Project', right_on=AX_PROJECT_COL,
        how='left', indicator=True,
    )

    merged['AX Project'] = merged['Project']

    # 매칭상태 판정
    def _status(row):
        if row['_merge'] == 'left_only':
            return 'NOAH에 없음'
        ax_amt = row['AX'] if pd.notna(row['AX']) else 0
        noah_amt = row['NOAH_금액_KRW'] if pd.notna(row['NOAH_금액_KRW']) else 0
        return '일치' if abs(ax_amt - noah_amt) < 1 else '불일치'

    merged['매칭상태'] = merged.apply(_status, axis=1)

    # Customer: AX 파일 우선
    if 'Customer_x' in merged.columns:
        merged['Customer'] = merged['Customer_x'].fillna(merged['Customer_y'])
    else:
        merged['Customer'] = merged.get('Customer', '')

    # 차이 계산
    merged['AX_금액'] = merged['AX'].fillna(0)
    merged['NOAH_금액(KRW)'] = merged['NOAH_금액_KRW'].fillna(0)
    merged['차이'] = merged['AX_금액'] - merged['NOAH_금액(KRW)']

    # --- 환율차이 판별: 불일치 + 외화 건에 대해 대사월 환율로 재계산 ---
    merged['재계산_KRW'] = np.nan
    merged['대사월_환율'] = np.nan

    if fx_rates is not None and recon_period_col is not None:
        for idx, row in merged.iterrows():
            if row['매칭상태'] != '불일치':
                continue
            cur = row.get('Currency')
            fx_amt = row.get('외화금액')
            if pd.isna(cur) or cur == 'KRW' or pd.isna(fx_amt) or fx_amt == 0:
                continue
            if cur not in fx_rates.index or recon_period_col not in fx_rates.columns:
                continue

            rate = fx_rates.loc[cur, recon_period_col]
            recalc = fx_amt * rate
            merged.at[idx, '대사월_환율'] = rate
            merged.at[idx, '재계산_KRW'] = recalc

            ax_amt = row['AX_금액']
            if abs(recalc - ax_amt) < FX_DIFF_THRESHOLD:
                merged.at[idx, '매칭상태'] = '일치(환율차이)'

    # Summary
    summary_cols = ['AX Project', 'Customer', 'AX_금액', 'NOAH_금액(KRW)', '차이',
                    'Currency', '외화금액', '환율', '대사월_환율', '재계산_KRW', '매칭상태']
    summary = merged[summary_cols].sort_values('AX Project').reset_index(drop=True)

    return summary


def write_output(
    summary: pd.DataFrame,
    detail: pd.DataFrame,
    output_file: Path,
) -> None:
    """대사 결과 Excel 파일 출력"""
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    def _add_table(writer, sheet_name: str, display_name: str) -> None:
        ws = writer.sheets[sheet_name]
        if ws.max_row < 2:
            return
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tbl = Table(displayName=display_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)

    # 범례
    legend = pd.DataFrame([
        ['일치', 'AX = NOAH DN (차이 < 1원)'],
        ['일치(환율차이)', '외화 × 대사월 환율 = AX (DN 등록월 vs 대사월 환율 차이)'],
        ['불일치', 'AX ≠ NOAH DN (차이 ≥ 1원, 환율차이도 아님)'],
        ['NOAH에 없음', 'AX에 있지만 NOAH DN에 매칭 안됨'],
    ], columns=['매칭상태', '설명'])

    # 상세 시트용 컬럼 정리
    detail_cols = [c for c in [
        AX_PROJECT_COL, 'DN_ID', 'SO_ID', 'Line item', 'Customer name',
        'Item', 'Qty', 'Unit Price', 'Total Sales',
        'Currency', 'Total Sales KRW', '출고일', '구분',
    ] if c in detail.columns]
    detail_out = detail[detail_cols].copy()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name='대사', index=False)
        _add_table(writer, '대사', '대사')

        if len(detail_out) > 0:
            detail_out.to_excel(writer, sheet_name='상세', index=False)
            _add_table(writer, '상세', '상세')

        legend.to_excel(writer, sheet_name='범례', index=False)
        _add_table(writer, '범례', '범례')


def print_summary(summary: pd.DataFrame) -> None:
    """대사 결과 요약 콘솔 출력"""
    print()
    print("SO 매출대사 결과")
    print("=" * 60)

    total = len(summary)
    matched = len(summary[summary['매칭상태'] == '일치'])
    fx_diff = len(summary[summary['매칭상태'] == '일치(환율차이)'])
    mismatched = len(summary[summary['매칭상태'] == '불일치'])
    noah_missing = len(summary[summary['매칭상태'] == 'NOAH에 없음'])

    print(f"  전체: {total}건")
    print(f"    일치:          {matched}")
    if fx_diff > 0:
        print(f"    일치(환율차이): {fx_diff}")
    if mismatched > 0:
        print(f"    불일치:        {mismatched}")
    if noah_missing > 0:
        print(f"    NOAH에 없음:   {noah_missing}")

    # 환율차이 상세
    fx_diff_rows = summary[summary['매칭상태'] == '일치(환율차이)']
    if len(fx_diff_rows) > 0:
        print()
        print("  환율차이 내역 (대사월 환율 적용 시 일치):")
        print(f"    {'AX Project':<16} {'AX_금액':>14} {'NOAH_금액':>14} {'차이':>14}  {'DN환율':>10} → {'대사월환율':>10}")
        print("    " + "-" * 96)
        for _, row in fx_diff_rows.iterrows():
            dn_rate = f"{row['환율']:>10,.2f}" if pd.notna(row['환율']) else ''
            ax_rate = f"{row['대사월_환율']:>10,.2f}" if pd.notna(row['대사월_환율']) else ''
            print(f"    {row['AX Project']:<16} {row['AX_금액']:>14,.0f} "
                  f"{row['NOAH_금액(KRW)']:>14,.0f} {row['차이']:>14,.0f}  {dn_rate} → {ax_rate}")

    # 불일치 상세
    mismatch_rows = summary[summary['매칭상태'] == '불일치']
    if len(mismatch_rows) > 0:
        print()
        print("  불일치 내역:")
        print(f"    {'AX Project':<16} {'AX_금액':>14} {'NOAH_금액':>14} {'차이':>14}  Currency")
        print("    " + "-" * 76)
        for _, row in mismatch_rows.iterrows():
            cur = row['Currency'] if pd.notna(row['Currency']) else ''
            fx = f"  {row['외화금액']:>12,.2f}" if pd.notna(row['외화금액']) else ''
            print(f"    {row['AX Project']:<16} {row['AX_금액']:>14,.0f} "
                  f"{row['NOAH_금액(KRW)']:>14,.0f} {row['차이']:>14,.0f}  {cur}{fx}")

    # NOAH에 없음 상세
    noah_missing_rows = summary[summary['매칭상태'] == 'NOAH에 없음']
    if len(noah_missing_rows) > 0:
        print()
        print("  NOAH에 없음:")
        for _, row in noah_missing_rows.iterrows():
            print(f"    {row['AX Project']:<16} AX {row['AX_금액']:>14,.0f}  {row['Customer']}")

    # 합계
    ax_total = summary['AX_금액'].sum()
    noah_total = summary['NOAH_금액(KRW)'].sum()
    print()
    print(f"  AX 합계:     {ax_total:>14,.0f}")
    print(f"  NOAH 합계:   {noah_total:>14,.0f}")
    print(f"  차이 합계:   {ax_total - noah_total:>14,.0f}")


def create_argument_parser() -> argparse.ArgumentParser:
    """CLI 인자 파서 생성"""
    parser = argparse.ArgumentParser(
        prog='reconcile_so',
        description='SO 매출대사 — AX ERP 매출 vs NOAH DN 매출 금액 비교',
        epilog='예시: python reconcile_so.py P03',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        'period',
        help='대사 월 코드 (예: P03, P04)',
    )

    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='상세 로그 출력',
    )

    return parser


def main() -> int:
    """메인 함수"""
    parser = create_argument_parser()
    args = parser.parse_args()

    setup_logging(verbose=args.verbose)
    period = args.period.upper()

    # 1. 파일 찾기
    ax_file = find_ax_sales_file(period)
    if not ax_file:
        print(f"[오류] AX_Sales 파일을 찾을 수 없습니다 (so_reconciliation/{period}/AX_Sales*)")
        return 1

    if not NOAH_SO_PO_DN_FILE.exists():
        print(f"[오류] NOAH_SO_PO_DN.xlsx를 찾을 수 없습니다: {NOAH_SO_PO_DN_FILE}")
        return 1

    print(f"AX Sales:  {ax_file.name}")
    print(f"NOAH DN:   {NOAH_SO_PO_DN_FILE.name}")

    # 2. 데이터 로드
    #    P03 → '2026-03' (FX 시트 컬럼에서 연도 파악)
    try:
        ax_sales = load_ax_sales(ax_file)
        fx_rates = load_fx_rates()
        recon_period_col = period_code_to_col(period, list(fx_rates.columns))
    except Exception as e:
        print(f"[오류] 데이터 로드 실패: {e}")
        return 1

    if not recon_period_col:
        print(f"[오류] FX 시트에 {period} 해당 환율 없음")
        return 1

    print(f"FX 환율:   {recon_period_col} 적용")

    try:
        df_dn = load_noah_dn(recon_period_col)
    except Exception as e:
        print(f"[오류] DN 로드 실패: {e}")
        return 1

    print(f"DN 필터:   출고일 {recon_period_col} ({len(df_dn)}건)")

    # 3. NOAH DN 집계
    noah_agg = aggregate_noah_dn(df_dn)

    # 4. 대사 결과 생성 (AX Sales 기준 left join)
    summary = build_reconciliation(ax_sales, noah_agg, fx_rates, recon_period_col)

    # 5. 상세 시트: 대사 대상 프로젝트의 DN 라인 상세
    all_projects = set(summary['AX Project'].dropna())
    detail = df_dn[df_dn[AX_PROJECT_COL].isin(all_projects)].copy()
    sort_cols = [c for c in [AX_PROJECT_COL, 'DN_ID', 'Line item'] if c in detail.columns]
    detail = detail.sort_values(sort_cols).reset_index(drop=True)

    # 6. Excel 출력
    period_dir = resolve_period_dir(RECON_DIR, period) or (RECON_DIR / period)
    output_file = period_dir / f"대사결과_SO_{period}.xlsx"
    write_output(summary, detail, output_file)
    print(f"\n출력: {output_file}")

    # 7. 콘솔 요약
    print_summary(summary)

    return 0


if __name__ == "__main__":
    sys.exit(main())
