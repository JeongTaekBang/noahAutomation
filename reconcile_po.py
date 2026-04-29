#!/usr/bin/env python
"""
PO 매입대사 (Reconciliation)
============================

공장 출고 금액(2026리스트_RCK)과 회계팀 GRN 금액(Noah purchase_GRN)을
AX PO 기준으로 비교하여 금액 일치 여부를 확인합니다.

사용법:
    python reconcile_po.py P03           # 3월 대사
    python reconcile_po.py P03 -v        # 상세 로그
"""

from __future__ import annotations

import argparse
import logging
import sys
import warnings
from pathlib import Path

import pandas as pd

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

from po_generator.config import NOAH_SO_PO_DN_FILE, BASE_DIR
from po_generator.logging_config import setup_logging
from po_generator.recon_paths import resolve_period_dir

logger = logging.getLogger(__name__)

RECON_DIR = BASE_DIR / "po_reconciliation"

# PO_국내: columns 0-15 (PO_ID ~ Status)
PO_DOMESTIC_COLS = [
    'PO_ID', 'SO_ID', 'Order type', 'Period', 'NOAH O.C No.', 'AX PO',
    'Customer name', 'Customer PO', 'Item name', 'Line item',
    'Item qty', 'ICO Unit', 'Total ICO', '공장 발주 날짜', '공장 EXW date', 'Status',
]

# PO_해외: columns 0-14 (PO_ID ~ Status, no "Order type")
PO_EXPORT_COLS = [
    'PO_ID', 'SO_ID', 'Period', 'NOAH O.C No.', 'AX PO',
    'Customer name', 'Customer PO', 'Item name', 'Line item',
    'Item qty', 'ICO Unit', 'Total ICO', '공장 발주 날짜', '공장 EXW date', 'Status',
]

GRN_USE_COLS = [
    'Item number', 'Item name', 'Purchase order', 'Quantity', 'Unit',
    'Date received', 'Voucher physical', 'Cost amount physical',
    'Date financial', 'Voucher financial', 'Cost amount financial',
    'Vendor name',
]


def find_file(period_code: str, keyword: str) -> Path | None:
    """period 디렉터리(플랫 또는 연도 중첩)에서 keyword가 포함된 Excel 파일 찾기"""
    period_dir = resolve_period_dir(RECON_DIR, period_code)
    if period_dir is None:
        return None
    for f in period_dir.iterdir():
        if f.suffix == '.xlsx' and not f.name.startswith('~'):
            if keyword.upper() in f.name.upper():
                return f
    return None


def load_internal_po() -> tuple[pd.DataFrame, pd.DataFrame]:
    """NOAH_SO_PO_DN.xlsx에서 PO_국내/PO_해외 읽기 (PO_ID ~ Status)

    Returns:
        (AX PO 있는 PO, 전체 PO — Cancelled 제외)
    """
    xf = pd.ExcelFile(NOAH_SO_PO_DN_FILE)

    # PO_국내
    df_dom = pd.read_excel(xf, 'PO_국내')
    dom_cols = [c for c in PO_DOMESTIC_COLS if c in df_dom.columns]
    df_dom = df_dom[dom_cols].copy()
    df_dom['구분'] = '국내'

    # PO_해외
    df_exp = pd.read_excel(xf, 'PO_해외')
    exp_cols = [c for c in PO_EXPORT_COLS if c in df_exp.columns]
    df_exp = df_exp[exp_cols].copy()
    df_exp['구분'] = '해외'

    # 통합
    df_all = pd.concat([df_dom, df_exp], ignore_index=True)

    # Cancelled 제외
    if 'Status' in df_all.columns:
        cancelled = df_all['Status'].astype(str).str.lower().str.contains('cancel', na=False)
        n_cancelled = cancelled.sum()
        if n_cancelled > 0:
            logger.debug("Cancelled 제외: %d건", n_cancelled)
            df_all = df_all[~cancelled].copy()

    # AX PO 있는 건
    has_ax = df_all['AX PO'].notna() & (df_all['AX PO'] != '')
    df = df_all[has_ax].copy()
    df['AX PO'] = df['AX PO'].astype(str).str.strip()

    logger.debug("내부 PO 로드: %d건 (AX PO 있음), %d건 (전체)",
                 len(df), len(df_all))
    return df, df_all


def load_delivery(delivery_file: Path) -> pd.DataFrame:
    """공장 출고 리스트 Delivery 시트 읽기"""
    df = pd.read_excel(delivery_file, sheet_name='Delivery')
    df = df[df['RCK ODER'].notna()].copy()
    df['RCK ODER'] = df['RCK ODER'].astype(str).str.strip()
    logger.debug("출고 리스트 로드: %d건 (%s)", len(df), delivery_file.name)
    return df


def load_grn(grn_file: Path) -> pd.DataFrame:
    """회계 GRN 파일 읽기 — 'Purchase order' 컬럼이 있는 시트를 자동 선택"""
    xf = pd.ExcelFile(grn_file)
    target_sheet = None
    for sn in xf.sheet_names:
        head = pd.read_excel(xf, sheet_name=sn, nrows=0)
        if 'Purchase order' in head.columns:
            target_sheet = sn
            break
    if target_sheet is None:
        raise ValueError(
            f"GRN 파일에서 'Purchase order' 컬럼이 있는 시트를 찾을 수 없습니다 "
            f"(시트 목록: {xf.sheet_names})"
        )
    df = pd.read_excel(xf, sheet_name=target_sheet)
    use_cols = [c for c in GRN_USE_COLS if c in df.columns]
    df = df[use_cols].copy()
    df = df[df['Purchase order'].notna()].copy()
    df['Purchase order'] = df['Purchase order'].astype(str).str.strip()
    logger.debug("GRN 로드: %d건 (%s, 시트=%s)", len(df), grn_file.name, target_sheet)
    return df


def build_po_mapping(df_po: pd.DataFrame) -> pd.DataFrame:
    """PO_ID → AX PO 매핑 테이블 (1:N 지원)"""
    mapping = df_po[['PO_ID', 'AX PO']].drop_duplicates()
    mapping = mapping[mapping['AX PO'].notna()].copy()
    mapping['PO_ID'] = mapping['PO_ID'].astype(str).str.strip()
    mapping['AX PO'] = mapping['AX PO'].astype(str).str.strip()
    mapping = mapping[mapping['AX PO'] != 'nan']
    return mapping


def resolve_ax_po(delivery: pd.DataFrame, po_mapping: pd.DataFrame) -> pd.DataFrame:
    """Delivery의 RCK ODER를 AX PO로 통일 (1:N → 행 복제)"""
    is_nd = delivery['RCK ODER'].str.startswith(('ND-', 'NO-'))
    df_nd = delivery[is_nd].merge(
        po_mapping, left_on='RCK ODER', right_on='PO_ID', how='left')
    df_nd['AX PO'] = df_nd['AX PO'].fillna('')

    df_p = delivery[~is_nd].copy()
    df_p['AX PO'] = df_p['RCK ODER']

    result = pd.concat([df_nd, df_p], ignore_index=True)

    unmapped = result[result['AX PO'] == '']
    if len(unmapped) > 0:
        logger.warning("AX PO 매핑 실패 %d건: %s",
                       len(unmapped), unmapped['RCK ODER'].tolist()[:5])
    return result


def export_delivery_ax_po(delivery: pd.DataFrame, output_file: Path,
                          df_po_export: pd.DataFrame | None = None) -> None:
    """Delivery 시트 + AX PO 매핑 결과를 별도 파일로 출력 (회계팀 GRN 대사용)

    Args:
        delivery: resolve_ax_po() 적용된 국내 Delivery DataFrame
        output_file: 출력 경로
        df_po_export: PO_해외 Invoiced 데이터 (없으면 국내만 출력)
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    def _add_tbl(writer, sheet_name: str, display_name: str) -> None:
        ws = writer.sheets[sheet_name]
        if ws.max_row >= 2:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tbl = Table(displayName=display_name, ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tbl)

    # --- 국내: Delivery + AX PO ---
    df = delivery.copy()
    drop_cols = [c for c in ['PO_ID'] if c in df.columns]
    if drop_cols:
        df = df.drop(columns=drop_cols)

    # 1:N 매핑 → AX PO 콤마로 합치기
    orig_cols = [c for c in df.columns if c != 'AX PO']
    df = (df.groupby(orig_cols, sort=False, dropna=False)
            .agg({'AX PO': lambda x: ', '.join(sorted(set(x.dropna().astype(str))))})
            .reset_index())

    # AX PO를 RCK ODER 바로 뒤에 배치
    cols = list(df.columns)
    if 'AX PO' in cols and 'RCK ODER' in cols:
        cols.remove('AX PO')
        idx = cols.index('RCK ODER') + 1
        cols.insert(idx, 'AX PO')
        df = df[cols]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='국내_Delivery', index=False)
        _add_tbl(writer, '국내_Delivery', '국내_Delivery')

        # --- 해외: PO_해외 Invoiced → 같은 구조로 매핑 ---
        if df_po_export is not None and len(df_po_export) > 0:
            exp = df_po_export.copy()
            # PO_ID별 집계: AX PO 콤마, 금액 합산
            exp_agg = (exp.groupby('PO_ID', sort=False)
                          .agg(
                              AX_PO=('AX PO', lambda x: ', '.join(
                                  sorted(set(x.dropna().astype(str).str.strip()) - {'', 'nan'}))),
                              SO_ID=('SO_ID', 'first'),
                              Customer=('Customer name', 'first'),
                              계산서금액=('Total ICO', 'sum'),
                          )
                          .reset_index()
                          .rename(columns={'PO_ID': 'RCK ODER', 'AX_PO': 'AX PO'}))
            exp_agg = exp_agg[['RCK ODER', 'AX PO', 'SO_ID', 'Customer', '계산서금액']]

            exp_agg.to_excel(writer, sheet_name='해외_PO', index=False)
            _add_tbl(writer, '해외_PO', '해외_PO')


def _agg_delivery(delivery: pd.DataFrame, name: str) -> pd.DataFrame:
    """출고 리스트를 AX PO별로 집계"""
    agg = (delivery.groupby('AX PO')
           .agg(
               Type=('Type', 'first'),
               Customer=('Customer', 'first'),
               SECTOR=('SECTOR', 'first'),
               RCK_ODER=('RCK ODER', lambda x: ', '.join(sorted(set(x)))),
               SO_ID=('SO_ID', lambda x: ', '.join(sorted(set(x.dropna().astype(str)))) if x.notna().any() else ''),
               건수=('AX PO', 'size'),
               금액=('계산서금액', 'sum'),
           )
           .reset_index())
    agg['수량'] = agg['건수']
    agg['금액_financial'] = 0
    agg['소스'] = name
    return agg


def _agg_po(df_po: pd.DataFrame, name: str) -> pd.DataFrame:
    """내부 PO를 AX PO별로 집계"""
    agg = (df_po.groupby('AX PO')
           .agg(
               Type=('구분', 'first'),
               Customer=('Customer name', 'first'),
               RCK_ODER=('PO_ID', lambda x: ', '.join(sorted(set(x.astype(str))))),
               SO_ID=('SO_ID', lambda x: ', '.join(sorted(set(x.dropna().astype(str))))),
               건수=('AX PO', 'size'),
               수량=('Item qty', 'sum'),
               금액=('Total ICO', 'sum'),
           )
           .reset_index())
    agg['SECTOR'] = ''
    agg['금액_financial'] = 0
    agg['소스'] = name
    return agg


def _agg_grn(grn: pd.DataFrame, name: str) -> pd.DataFrame:
    """GRN을 AX PO별로 집계"""
    agg = (grn.groupby('Purchase order')
           .agg(
               건수=('Purchase order', 'size'),
               수량=('Quantity', 'sum'),
               금액=('Cost amount physical', 'sum'),
               금액_financial=('Cost amount financial', 'sum'),
           )
           .reset_index()
           .rename(columns={'Purchase order': 'AX PO'}))
    agg['Type'] = ''
    agg['Customer'] = ''
    agg['SECTOR'] = ''
    agg['RCK_ODER'] = ''
    agg['SO_ID'] = ''
    agg['소스'] = name
    return agg


OUT_COLS = ['AX PO', '소스', 'Type', 'Customer', 'SECTOR', 'RCK_ODER', 'SO_ID',
            '건수', '수량', '금액', '금액_financial']


def build_raw_data(delivery: pd.DataFrame, df_po: pd.DataFrame,
                   grn: pd.DataFrame,
                   delivery_name: str = '', po_name: str = '',
                   grn_name: str = '') -> tuple[pd.DataFrame, pd.DataFrame]:
    """GRN 기준 long format raw_data + GRN 미포함 건 반환"""
    del_agg = _agg_delivery(delivery, delivery_name)
    po_agg = _agg_po(df_po, po_name)
    grn_agg = _agg_grn(grn, grn_name)

    grn_set = set(grn_agg['AX PO'])
    del_set = set(del_agg['AX PO'])
    po_set = set(po_agg['AX PO'])

    # --- 매칭 상태: PO vs GRN 기준 (GRN 행에만 표시) ---
    status_map = {}
    for ax in grn_set:
        has_del = ax in del_set
        has_po = ax in po_set
        grn_amt = grn_agg.loc[grn_agg['AX PO'] == ax, '금액'].iloc[0]

        if has_po:
            po_amt = po_agg.loc[po_agg['AX PO'] == ax, '금액'].iloc[0]
            po_type = po_agg.loc[po_agg['AX PO'] == ax, 'Type'].iloc[0]
            matched = abs(po_amt - grn_amt) < 1
            if po_type == '해외':
                status_map[ax] = '일치(해외)' if matched else '불일치(해외)'
            else:
                status_map[ax] = '일치' if matched else '불일치'
        elif has_del:
            del_type = del_agg.loc[del_agg['AX PO'] == ax, 'Type'].iloc[0]
            del_amt = del_agg.loc[del_agg['AX PO'] == ax, '금액'].iloc[0]
            if del_type in ('Service', 'YTC'):
                if abs(del_amt - grn_amt) < 1:
                    status_map[ax] = '일치(서비스)'
                else:
                    status_map[ax] = '불일치(서비스)'
            else:
                status_map[ax] = 'AX PO 누락'
        else:
            status_map[ax] = '서비스/기타'

    # --- raw_data: GRN에 있는 AX PO만 ---
    grn_rows = grn_agg[OUT_COLS]
    del_rows = del_agg[del_agg['AX PO'].isin(grn_set)][OUT_COLS]
    po_rows = po_agg[po_agg['AX PO'].isin(grn_set)][OUT_COLS]

    result = pd.concat([grn_rows, del_rows, po_rows], ignore_index=True)

    # 매칭상태: GRN 행에만
    result['매칭상태'] = ''
    grn_mask = result['소스'] == grn_name
    result.loc[grn_mask, '매칭상태'] = result.loc[grn_mask, 'AX PO'].map(status_map)

    # GRN 첫 번째로 정렬 (GRN name이 알파벳상 먼저 오도록 sort key)
    result['_sort'] = result['소스'].apply(
        lambda s: 0 if s == grn_name else (1 if s == delivery_name else 2))
    result = result.sort_values(['AX PO', '_sort']).drop(columns='_sort').reset_index(drop=True)

    for col in ['건수', '수량']:
        result[col] = result[col].fillna(0).astype(int)

    # --- GRN 미포함: 출고/PO에는 있지만 당월 GRN에 없는 건 (AX에 GRN 처리 필요) ---
    not_in_grn_ax = (del_set | po_set) - grn_set
    if not_in_grn_ax:
        del_miss = del_agg[del_agg['AX PO'].isin(not_in_grn_ax)][OUT_COLS]
        po_miss = po_agg[po_agg['AX PO'].isin(not_in_grn_ax)][OUT_COLS]
        missing = pd.concat([del_miss, po_miss], ignore_index=True)
        missing['매칭상태'] = 'GRN 미포함'
        missing = missing.sort_values(['AX PO', '소스']).reset_index(drop=True)
        for col in ['건수', '수량']:
            missing[col] = missing[col].fillna(0).astype(int)
    else:
        missing = pd.DataFrame(columns=result.columns)

    return result, missing



def print_summary_table(raw_data: pd.DataFrame, missing: pd.DataFrame,
                        delivery_name: str, grn_name: str) -> None:
    """대사 결과 요약 출력"""
    # GRN 행에서 매칭상태 추출
    status = raw_data[raw_data['매칭상태'] != ''][['AX PO', '매칭상태']].drop_duplicates()

    print()
    print("PO 매입대사 결과")
    print("=" * 60)

    total = len(status)
    matched = len(status[status['매칭상태'] == '일치'])
    matched_exp = len(status[status['매칭상태'] == '일치(해외)'])
    mismatched = len(status[status['매칭상태'] == '불일치'])
    mismatched_exp = len(status[status['매칭상태'] == '불일치(해외)'])
    matched_svc = len(status[status['매칭상태'] == '일치(서비스)'])
    mismatched_svc = len(status[status['매칭상태'] == '불일치(서비스)'])
    ax_missing = len(status[status['매칭상태'] == 'AX PO 누락'])
    service = len(status[status['매칭상태'] == '서비스/기타'])
    grn_not_included = missing['AX PO'].nunique() if len(missing) > 0 else 0

    print(f"  [GRN 기준] {total}건")
    print(f"    일치:          {matched}")
    print(f"    일치(해외):    {matched_exp}")
    if mismatched > 0:
        print(f"    불일치:        {mismatched}")
    if mismatched_exp > 0:
        print(f"    불일치(해외):  {mismatched_exp}")
    print(f"    일치(서비스):   {matched_svc}")
    if mismatched_svc > 0:
        print(f"    불일치(서비스): {mismatched_svc}")
    if ax_missing > 0:
        print(f"    AX PO 누락:    {ax_missing} (PO시트에 AX PO 미입력)")
    if service > 0:
        print(f"    서비스/기타:   {service} (GRN에만 존재)")
    print()
    print(f"  [GRN 미포함] {grn_not_included}건 (별도 시트)")
    print(f"    출고/PO에는 있는데 당월 GRN에 없는 건 — AX에 GRN 처리 필요")

    # 불일치 상세
    if mismatched + mismatched_exp + mismatched_svc > 0:
        mismatch_ax = status[status['매칭상태'].isin(
            ['불일치', '불일치(해외)', '불일치(서비스)'])]['AX PO'].tolist()
        po_src = [s for s in raw_data['소스'].unique() if s not in (delivery_name, grn_name)]
        po_src = po_src[0] if po_src else ''
        print()
        print("불일치 내역:")
        print(f"  {'AX PO':<12} {'비교금액':>14} {'GRN금액':>14} {'차이':>14}")
        print("  " + "-" * 56)
        for ax in mismatch_ax:
            ax_status = status.loc[status['AX PO'] == ax, '매칭상태'].iloc[0]
            g_amt = raw_data.loc[(raw_data['AX PO'] == ax) &
                                 (raw_data['소스'] == grn_name), '금액'].iloc[0]
            if '서비스' in ax_status:
                ref = raw_data.loc[(raw_data['AX PO'] == ax) &
                                    (raw_data['소스'] == delivery_name), '금액']
            else:
                ref = raw_data.loc[(raw_data['AX PO'] == ax) &
                                    (raw_data['소스'] == po_src), '금액']
            r_amt = ref.iloc[0] if len(ref) > 0 else 0
            print(f"  {ax:<12} {r_amt:>14,.0f} {g_amt:>14,.0f} {r_amt - g_amt:>14,.0f}")

    # 소스별 합계
    print()
    for src in raw_data['소스'].unique():
        src_total = raw_data.loc[raw_data['소스'] == src, '금액'].sum()
        print(f"  {src}:  {src_total:>14,.0f}")


def create_argument_parser() -> argparse.ArgumentParser:
    """CLI 인자 파서 생성"""
    parser = argparse.ArgumentParser(
        prog='reconcile_po',
        description='PO 매입대사 — 공장 출고 vs 회계 GRN 금액 비교',
        epilog='예시: python reconcile_po.py P03',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        'period',
        help='대사 월 코드 (예: P03, P04)',
    )

    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='상세 로그 출력',
    )

    return parser


def main() -> int:
    """메인 함수"""
    parser = create_argument_parser()
    args = parser.parse_args()

    setup_logging(verbose=args.verbose)
    period = args.period.upper()

    # 1. 파일 찾기
    grn_file = find_file(period, 'GRN')
    if not grn_file:
        print(f"[오류] GRN 파일을 찾을 수 없습니다 (po_reconciliation/*{period}*GRN*)")
        return 1

    delivery_file = find_file(period, '리스트')
    if not delivery_file:
        # 영문 fallback
        delivery_file = find_file(period, 'RCK')
    if not delivery_file:
        print(f"[오류] 출고 리스트 파일을 찾을 수 없습니다 (po_reconciliation/*{period}*리스트*)")
        return 1

    if not NOAH_SO_PO_DN_FILE.exists():
        print(f"[오류] NOAH_SO_PO_DN.xlsx를 찾을 수 없습니다: {NOAH_SO_PO_DN_FILE}")
        return 1

    period_dir = resolve_period_dir(RECON_DIR, period) or (RECON_DIR / period)

    print(f"GRN 파일:    {grn_file.name}")
    print(f"출고 리스트: {delivery_file.name}")
    print(f"내부 PO:     {NOAH_SO_PO_DN_FILE.name}")

    # 2. 데이터 로드
    try:
        df_po, df_po_all = load_internal_po()
        delivery = load_delivery(delivery_file)
        grn = load_grn(grn_file)
    except Exception as e:
        print(f"[오류] 데이터 로드 실패: {e}")
        return 1

    # 3. PO_ID → AX PO 매핑 (전체 PO로)
    po_mapping = build_po_mapping(df_po)
    delivery = resolve_ax_po(delivery, po_mapping)

    # 3-1. Delivery(국내) + PO_해외 → AX PO 매핑 파일 출력 (회계팀 GRN 대사용)
    invoiced_status = f"Invoiced {period}"
    df_po_export = df_po_all[
        (df_po_all['구분'] == '해외') &
        (df_po_all['Status'] == invoiced_status)
    ].copy()
    ax_po_file = period_dir / f"AX_PO_매핑_{period}.xlsx"
    export_delivery_ax_po(delivery, ax_po_file, df_po_export=df_po_export)
    print(f"AX PO 매핑: {ax_po_file.name}")

    delivery = delivery[delivery['AX PO'] != ''].copy()

    # 4. PO를 당월 Invoiced만 필터링
    if 'Status' in df_po.columns:
        df_po_period = df_po[df_po['Status'] == invoiced_status].copy()
    else:
        df_po_period = df_po

    # 4-1. AX PO 미입력 건 (Invoiced인데 AX PO 없음)
    ax_empty = df_po_all[
        (df_po_all['Status'] == invoiced_status) &
        (df_po_all['AX PO'].isna() | (df_po_all['AX PO'] == ''))
    ].copy()
    show_cols = [c for c in ['PO_ID', 'SO_ID', 'Customer name', 'Item name',
                             'Item qty', 'Total ICO', '구분'] if c in ax_empty.columns]
    ax_empty = ax_empty[show_cols]

    # 5. raw_data 생성
    raw_data, missing = build_raw_data(
        delivery, df_po_period, grn,
        delivery_name=delivery_file.stem,
        po_name=NOAH_SO_PO_DN_FILE.stem,
        grn_name=grn_file.stem,
    )

    # 5. Excel 출력
    output_file = period_dir / f"대사결과_{period}.xlsx"
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    def _add_table(writer, sheet_name: str, display_name: str) -> None:
        ws = writer.sheets[sheet_name]
        if ws.max_row < 2:
            return
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tbl = Table(displayName=display_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)

    def _append_totals_block(
        ws, summary_df: pd.DataFrame,
        ax_empty_df: pd.DataFrame, missing_summary_df: pd.DataFrame,
    ) -> None:
        """대사 시트 표 아래에 합계 블록 추가."""
        grn_total = float(summary_df['GRN_금액'].sum())
        cmp_total = float(summary_df['비교_금액'].sum())
        gap = grn_total - cmp_total
        miss_total = (float(missing_summary_df['매입금액'].sum())
                      if '매입금액' in missing_summary_df.columns else 0.0)
        ax_empty_total = float(ax_empty_df['Total ICO'].sum()) if (
            len(ax_empty_df) and 'Total ICO' in ax_empty_df.columns
        ) else 0.0

        bold = Font(bold=True)
        rows = [
            ('GRN 합계',         None,           grn_total,      '전체 GRN (raw 합계와 일치)'),
            ('비교가능 합계',    cmp_total,      None,           'PO/출고에 매칭된 금액'),
            ('비교불가',         None,           gap,            '서비스/기타 — GRN만 존재'),
            ('GRN 미포함 합계',  miss_total,     None,           '별도 시트 — AX에 GRN 처리 필요'),
            ('AX PO 미입력',     ax_empty_total, None,           '별도 시트 — Invoiced인데 AX PO 없음'),
        ]
        start = ws.max_row + 2
        for i, (label, b_val, c_val, note) in enumerate(rows):
            r = start + i
            cell = ws.cell(r, 1, label)
            cell.font = bold
            if b_val is not None:
                c = ws.cell(r, 2, b_val)
                c.number_format = '#,##0'
            if c_val is not None:
                c = ws.cell(r, 3, c_val)
                c.number_format = '#,##0'
            if note:
                ws.cell(r, 4, note)

    # 매칭상태 범례
    legend = pd.DataFrame([
        ['일치', 'O', '-', 'O', 'PO = GRN (국내)'],
        ['불일치', 'O', '-', 'O', 'PO ≠ GRN (국내)'],
        ['일치(해외)', 'O', '-', 'O', 'PO = GRN (해외)'],
        ['불일치(해외)', 'O', '-', 'O', 'PO ≠ GRN (해외)'],
        ['일치(서비스)', 'O', 'O', 'X', '출고 = GRN (서비스/YTC)'],
        ['불일치(서비스)', 'O', 'O', 'X', '출고 ≠ GRN (서비스/YTC)'],
        ['AX PO 누락', 'O', 'O', 'X', 'Product인데 PO시트에 AX PO 미입력'],
        ['서비스/기타', 'O', 'X', 'X', 'GRN에만 존재 (출고리스트에도 없음)'],
        ['GRN 미포함', 'X', '-', '-', '출고/PO에는 있는데 GRN에 없음 — AX에 GRN 처리 필요 (별도 시트, 출처별)'],
    ], columns=['매칭상태', 'GRN', '출고리스트', '내부PO', '설명'])

    # GRN 대사 요약: AX PO별 1행
    grn_name = grn_file.stem
    del_name = delivery_file.stem
    po_name = NOAH_SO_PO_DN_FILE.stem

    grn_rows = raw_data[raw_data['소스'] == grn_name][['AX PO', '금액', '매칭상태']].copy()
    grn_rows = grn_rows.rename(columns={'금액': 'GRN_금액'})

    # PO 금액
    po_rows = raw_data[raw_data['소스'] == po_name][['AX PO', '금액']].copy()
    po_rows = po_rows.rename(columns={'금액': 'PO_금액'})

    # 출고 금액 (서비스 비교용)
    del_rows = raw_data[raw_data['소스'] == del_name][['AX PO', '금액']].copy()
    del_rows = del_rows.rename(columns={'금액': '출고_금액'})

    summary = grn_rows.merge(po_rows, on='AX PO', how='left')
    summary = summary.merge(del_rows, on='AX PO', how='left')

    # 비교금액: PO 우선, 없으면 출고 (서비스)
    summary['비교_금액'] = summary['PO_금액'].fillna(summary['출고_금액'])
    summary = summary[['AX PO', '비교_금액', 'GRN_금액', '매칭상태']]
    summary = summary.sort_values('AX PO').reset_index(drop=True)

    # GRN 미포함 요약: AX PO + 매입금액 (PO 우선, 없으면 출고)
    if len(missing) > 0:
        pivoted = (missing.pivot_table(
                index='AX PO', columns='소스', values='금액',
                aggfunc='sum', fill_value=0)
            .reset_index())
        rename_map = {po_name: 'PO_금액', del_name: '출고_금액'}
        pivoted = pivoted.rename(columns=rename_map)
        for col in ('PO_금액', '출고_금액'):
            if col not in pivoted.columns:
                pivoted[col] = 0
        pivoted['매입금액'] = pivoted['PO_금액'].where(
            pivoted['PO_금액'] > 0, pivoted['출고_금액'])
        missing_summary = (pivoted[['AX PO', '매입금액']]
                           .sort_values('매입금액', ascending=False)
                           .reset_index(drop=True))
    else:
        missing_summary = pd.DataFrame(columns=['AX PO', '매입금액'])

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name='대사', index=False)
        _add_table(writer, '대사', '대사')
        _append_totals_block(writer.sheets['대사'], summary,
                              ax_empty, missing_summary)

        if len(missing_summary) > 0:
            missing_summary.to_excel(writer, sheet_name='GRN_미포함', index=False)
            _add_table(writer, 'GRN_미포함', 'GRN_미포함')

        raw_data.to_excel(writer, sheet_name='raw_data', index=False)
        _add_table(writer, 'raw_data', 'raw_data')

        if len(missing) > 0:
            missing.to_excel(writer, sheet_name='raw_data_미포함', index=False)
            _add_table(writer, 'raw_data_미포함', 'raw_data_미포함')

        if len(ax_empty) > 0:
            ax_empty.to_excel(writer, sheet_name='AX_PO_미입력', index=False)
            _add_table(writer, 'AX_PO_미입력', 'AX_PO_미입력')

        legend.to_excel(writer, sheet_name='범례', index=False)
        _add_table(writer, '범례', '범례')

    print(f"\n출력: {output_file}")

    # 6. 요약 출력
    print_summary_table(raw_data, missing, delivery_file.stem, grn_file.stem)

    if len(ax_empty) > 0:
        print(f"\n  [주의] AX PO 미입력 {len(ax_empty)}건 — AX_PO_미입력 시트 확인")
        for _, row in ax_empty.head(10).iterrows():
            po_id = row.get('PO_ID', '')
            so_id = row.get('SO_ID', '')
            cust = row.get('Customer name', '')
            print(f"    {po_id}  {so_id}  {cust}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
