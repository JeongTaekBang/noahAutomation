#!/usr/bin/env python
"""
Industry Code 대사 (Industry Code Reconciliation)
===================================================

1) Orderbook 파일의 빈 Industry code를
   NOAH_SO_PO_DN.xlsx (PO → SO) 매핑을 통해 채워서 새 파일로 출력.
   추가 컬럼: NOAH Sector, 매핑상태

2) SO시트 Sector 검증 — Industry code 마스터(Orderbook)의 Category와
   SO시트 Sector가 일치하는지 교차 검증 리포트 생성.

사용법:
    python reconcile_ind.py P03           # P03 대사
    python reconcile_ind.py P03 -v        # 상세 로그
"""

from __future__ import annotations

import argparse
import logging
import sys
import warnings
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

from po_generator.config import (
    NOAH_SO_PO_DN_FILE, BASE_DIR,
    SO_DOMESTIC_SHEET, SO_EXPORT_SHEET,
    PO_DOMESTIC_SHEET, PO_EXPORT_SHEET,
)
from po_generator.logging_config import setup_logging

logger = logging.getLogger(__name__)

RECON_DIR = BASE_DIR / "ind_code_reconciliation"

# NOAH_SO_PO_DN 컬럼
PO_OC_COL = 'NOAH O.C No.'
PO_SO_ID_COL = 'SO_ID'
SO_SECTOR_COL = 'Sector'
SO_IND_COL = 'Industry code'

# Industry code 마스터 → Sector 약칭 매핑
CATEGORY_TO_SECTOR: dict[str, str] = {
    'Oil & Gas': 'OG',
    'Water & Power': 'WAPO',
    'Chemical, Process & Industrial': 'CPI',
}


def find_orderbook_file(period_code: str | None = None) -> Path | None:
    """Orderbook 파일 찾기

    period_code가 주어지면 해당 폴더에서,
    없으면 전체 하위 폴더에서 가장 최근 파일 반환.
    """
    if period_code:
        period_dir = RECON_DIR / period_code.upper()
        if not period_dir.exists():
            return None
        for f in period_dir.iterdir():
            if f.suffix == '.xlsx' and not f.name.startswith('~'):
                if 'orderbook' in f.name.lower() and '결과' not in f.name:
                    return f
        return None

    # period 없으면 전체 검색 (마스터 용도)
    candidates = []
    for sub in RECON_DIR.iterdir():
        if not sub.is_dir():
            continue
        for f in sub.iterdir():
            if f.suffix == '.xlsx' and not f.name.startswith('~'):
                if 'orderbook' in f.name.lower() and '결과' not in f.name:
                    candidates.append(f)
    if not candidates:
        return None
    # 가장 최근 수정 파일
    return max(candidates, key=lambda p: p.stat().st_mtime)


def get_ob_sheet_name(period_code: str) -> str:
    """기간 코드로 Orderbook 시트명 생성 (P03 → P03_start)"""
    return f"{period_code.upper()}_start"


def build_mapping() -> tuple[dict[str, tuple[str, object, str | None]], set[str]]:
    """발주번호 → (SO_ID, Industry code, Sector) 매핑 딕셔너리 + 중복 O.C No. 집합"""
    # PO: NOAH O.C No. → SO_ID
    po_dom = pd.read_excel(
        NOAH_SO_PO_DN_FILE, sheet_name=PO_DOMESTIC_SHEET,
        usecols=[PO_OC_COL, PO_SO_ID_COL],
    )
    po_exp = pd.read_excel(
        NOAH_SO_PO_DN_FILE, sheet_name=PO_EXPORT_SHEET,
        usecols=[PO_OC_COL, PO_SO_ID_COL],
    )
    po_map = pd.concat([po_dom, po_exp], ignore_index=True)
    po_map = po_map.dropna(subset=[PO_OC_COL, PO_SO_ID_COL])
    po_map[PO_OC_COL] = po_map[PO_OC_COL].astype(str).str.strip()
    po_map[PO_SO_ID_COL] = po_map[PO_SO_ID_COL].astype(str).str.strip()
    # PO 중복 경고: 같은 NOAH O.C No.가 다른 SO_ID에 매핑된 경우
    po_dup_mask = po_map.duplicated(subset=[PO_OC_COL], keep=False)
    if po_dup_mask.any():
        po_dups = (po_map[po_dup_mask]
                   .groupby(PO_OC_COL)[PO_SO_ID_COL]
                   .apply(list)
                   .to_dict())
        logger.warning(
            "PO 중복 NOAH O.C No. %d건 (첫 번째만 사용): %s",
            len(po_dups), po_dups,
        )
    po_map = po_map.drop_duplicates(subset=[PO_OC_COL])

    # SO: SO_ID → (Industry code, Sector)
    so_dom = pd.read_excel(
        NOAH_SO_PO_DN_FILE, sheet_name=SO_DOMESTIC_SHEET,
        usecols=[PO_SO_ID_COL, SO_IND_COL, SO_SECTOR_COL],
    )
    so_exp = pd.read_excel(
        NOAH_SO_PO_DN_FILE, sheet_name=SO_EXPORT_SHEET,
        usecols=[PO_SO_ID_COL, SO_IND_COL, SO_SECTOR_COL],
    )
    so_map = pd.concat([so_dom, so_exp], ignore_index=True)
    so_map = so_map.dropna(subset=[PO_SO_ID_COL])
    so_map[PO_SO_ID_COL] = so_map[PO_SO_ID_COL].astype(str).str.strip()
    # SO 중복 경고: 같은 SO_ID에 다른 Industry code/Sector 값
    so_dup_mask = so_map.duplicated(subset=[PO_SO_ID_COL], keep=False)
    if so_dup_mask.any():
        so_dups = (so_map[so_dup_mask]
                   .groupby(PO_SO_ID_COL)[[SO_IND_COL, SO_SECTOR_COL]]
                   .apply(lambda g: g.values.tolist())
                   .to_dict())
        logger.warning(
            "SO 중복 SO_ID %d건 (첫 번째만 사용): %s",
            len(so_dups), so_dups,
        )
    so_map = so_map.drop_duplicates(subset=[PO_SO_ID_COL])

    # Join & build dict
    full = po_map.merge(so_map, on=PO_SO_ID_COL, how='left')
    logger.debug("매핑: %d건 (Industry code 있음: %d건)",
                 len(full), full[SO_IND_COL].notna().sum())

    result = {}
    for _, row in full.iterrows():
        oc = row[PO_OC_COL]
        so_id = row[PO_SO_ID_COL]
        ind = row[SO_IND_COL] if pd.notna(row[SO_IND_COL]) else None
        sector = row[SO_SECTOR_COL] if pd.notna(row[SO_SECTOR_COL]) else None
        result[oc] = (so_id, ind, sector)

    dup_oc_set = set(po_dups.keys()) if po_dup_mask.any() else set()
    return result, dup_oc_set


def fill_industry_code(
    ob: pd.DataFrame, mapping: dict, ind_col: str,
    dup_oc_set: set[str] | None = None,
) -> tuple[pd.DataFrame, int, int, int, int, int]:
    """Orderbook에 Industry code, Sector, 매핑상태 채우기"""
    if dup_oc_set is None:
        dup_oc_set = set()
    result = ob.copy()
    result['NOAH Sector'] = None
    result['매핑상태'] = None

    null_mask = result[ind_col].isna()
    total_null = int(null_mask.sum())
    filled = 0
    so_missing = 0
    po_missing = 0
    dup_review = 0

    for idx in result[null_mask].index:
        order_no = result.at[idx, '발주번호']
        if pd.isna(order_no):
            continue
        order_no = str(order_no).strip()

        if order_no in dup_oc_set:
            # 중복 O.C No. → 값 채우지 않고 검토 대상으로 분리
            result.at[idx, '매핑상태'] = '중복검토'
            dup_review += 1
        elif order_no in mapping:
            so_id, ind_code, sector = mapping[order_no]
            if ind_code is not None:
                result.at[idx, ind_col] = ind_code
                result.at[idx, 'NOAH Sector'] = sector
                result.at[idx, '매핑상태'] = '매칭'
                filled += 1
            else:
                result.at[idx, 'NOAH Sector'] = sector
                result.at[idx, '매핑상태'] = 'SO에 Industry code 없음'
                so_missing += 1
        else:
            result.at[idx, '매핑상태'] = 'PO에 발주번호 없음'
            po_missing += 1

    return result, total_null, filled, so_missing, po_missing, dup_review


# ──────────────────────────────────────────────
# Sector 검증
# ──────────────────────────────────────────────

def load_industry_code_master(ob_file: Path) -> dict:
    """Orderbook의 'Industry code' 시트에서 마스터 로드

    Returns:
        dict: {industry_code_str: (Category, expected_sector)}
    """
    master = pd.read_excel(ob_file, sheet_name='Industry code',
                           usecols=['Category', 'New Industry Code'])
    master = master.dropna(subset=['New Industry Code'])

    result = {}
    for _, row in master.iterrows():
        code = str(row['New Industry Code']).strip()
        category = str(row['Category']).strip()
        expected_sector = CATEGORY_TO_SECTOR.get(category)
        result[code] = (category, expected_sector)

    return result


def validate_so_sector(
    ob_file: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """SO시트 Sector vs Industry code 마스터 교차 검증

    Returns:
        (mismatches_df, summary_df)
    """
    master = load_industry_code_master(ob_file)

    # SO 로드 (국내 + 해외)
    so_cols = [PO_SO_ID_COL, 'Customer name', SO_SECTOR_COL, SO_IND_COL]
    so_dom = pd.read_excel(NOAH_SO_PO_DN_FILE, sheet_name=SO_DOMESTIC_SHEET)
    so_dom = so_dom[[c for c in so_cols if c in so_dom.columns]].copy()
    so_dom['구분'] = '국내'

    so_exp = pd.read_excel(NOAH_SO_PO_DN_FILE, sheet_name=SO_EXPORT_SHEET)
    so_exp = so_exp[[c for c in so_cols if c in so_exp.columns]].copy()
    so_exp['구분'] = '해외'

    so_all = pd.concat([so_dom, so_exp], ignore_index=True)
    so_all = so_all.dropna(subset=[SO_IND_COL, SO_SECTOR_COL])
    # SO_ID 기준 중복 제거 (라인아이템은 동일 Sector)
    so_all = so_all.drop_duplicates(subset=[PO_SO_ID_COL])

    # 마스터 조인
    so_all['ind_code_str'] = so_all[SO_IND_COL].apply(
        lambda x: str(int(x)) if isinstance(x, float) and x == int(x) else str(x).strip()
    )
    so_all['마스터 Category'] = so_all['ind_code_str'].map(
        lambda c: master[c][0] if c in master else None
    )
    so_all['기대 Sector'] = so_all['ind_code_str'].map(
        lambda c: master[c][1] if c in master else None
    )

    # 불일치 판정
    so_all['검증결과'] = so_all.apply(
        lambda r: '불일치' if pd.notna(r['기대 Sector']) and r[SO_SECTOR_COL] != r['기대 Sector']
        else ('마스터 없음' if pd.isna(r['마스터 Category']) else '일치'),
        axis=1,
    )

    # 불일치 건만 추출
    mismatches = so_all[so_all['검증결과'] == '불일치'].copy()
    mismatches = mismatches[[
        PO_SO_ID_COL, '구분', 'Customer name',
        SO_IND_COL, '마스터 Category', SO_SECTOR_COL, '기대 Sector',
    ]].sort_values([SO_SECTOR_COL, SO_IND_COL, PO_SO_ID_COL]).reset_index(drop=True)

    # 요약: Industry code × Sector 교차표 (건수)
    cross = so_all.groupby([SO_IND_COL, SO_SECTOR_COL, '마스터 Category', '기대 Sector', '검증결과']).size()
    summary = cross.reset_index(name='건수')
    summary = summary.sort_values([SO_IND_COL, SO_SECTOR_COL]).reset_index(drop=True)

    return mismatches, summary


# ──────────────────────────────────────────────
# 출력
# ──────────────────────────────────────────────

def _add_table(writer, sheet_name: str, display_name: str) -> None:
    """시트에 Excel 테이블 추가"""
    ws = writer.sheets[sheet_name]
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tbl = Table(displayName=display_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)


def write_ind_code_output(result: pd.DataFrame, output_file: Path) -> None:
    """Industry code 채움 결과 Excel 출력"""
    legend = pd.DataFrame([
        ['매칭', '발주번호 → PO → SO 체인으로 Industry code 찾아 채움'],
        ['중복검토', '같은 NOAH O.C No.가 복수 SO에 매핑 — 자동 채움 보류, 수동 확인 필요'],
        ['SO에 Industry code 없음', 'PO에서 SO_ID 찾았으나 SO시트에 Industry code 비어있음'],
        ['PO에 발주번호 없음', 'NOAH_SO_PO_DN PO시트에 해당 NOAH O.C No. 없음'],
    ], columns=['매칭상태', '설명'])

    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        result.to_excel(writer, sheet_name='결과', index=False)
        _add_table(writer, '결과', '결과')

        legend.to_excel(writer, sheet_name='범례', index=False)
        _add_table(writer, '범례', '범례')


def write_sector_output(
    mismatches: pd.DataFrame,
    sector_summary: pd.DataFrame,
    output_file: Path,
) -> None:
    """Sector 검증 결과 Excel 출력"""
    legend = pd.DataFrame([
        ['일치', 'SO Sector = Industry code 마스터의 기대 Sector'],
        ['불일치', 'SO Sector ≠ 기대 Sector (마스터 Category 기준)'],
        ['마스터 없음', 'Industry code가 마스터에 없음'],
    ], columns=['검증결과', '설명'])

    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        if len(mismatches) > 0:
            mismatches.to_excel(writer, sheet_name='불일치', index=False)
            _add_table(writer, '불일치', '불일치')

        sector_summary.to_excel(writer, sheet_name='검증요약', index=False)
        _add_table(writer, '검증요약', '검증요약')

        legend.to_excel(writer, sheet_name='범례', index=False)
        _add_table(writer, '범례', '범례')


def print_ind_summary(
    total_null: int, filled: int, so_missing: int, po_missing: int,
    dup_review: int = 0,
) -> None:
    """Industry code 채움 콘솔 요약"""
    print()
    print("Industry Code 채움 결과")
    print("=" * 60)
    print(f"  빈 Industry code: {total_null}행")
    print(f"    매칭 (채움):             {filled}")
    if dup_review > 0:
        print(f"    중복검토 (미채움):        {dup_review}")
    if so_missing > 0:
        print(f"    SO에 Industry code 없음: {so_missing}")
    if po_missing > 0:
        print(f"    PO에 발주번호 없음:       {po_missing}")
    print(f"  채움 후 남은 빈 행:        {total_null - filled}")


def print_sector_summary(
    mismatches: pd.DataFrame, sector_summary: pd.DataFrame,
) -> None:
    """Sector 검증 콘솔 요약"""
    mismatch_count = len(mismatches)
    match_count = int(sector_summary[sector_summary['검증결과'] == '일치']['건수'].sum())
    mismatch_total = int(sector_summary[sector_summary['검증결과'] == '불일치']['건수'].sum())

    print()
    print("SO Sector 검증")
    print("=" * 60)
    print(f"  검증 대상:   {match_count + mismatch_total}건 (SO_ID 기준)")
    print(f"    일치:      {match_count}")
    print(f"    불일치:    {mismatch_total} ({mismatch_count}건 SO_ID)")

    if mismatch_count > 0:
        print()
        print("  불일치 내역:")
        print(f"    {'SO_ID':<18} {'구분':<5} {'Ind.Code':>8}  {'현재Sector':<8} {'기대Sector':<8}  Customer")
        print("    " + "-" * 80)
        for _, r in mismatches.head(20).iterrows():
            ind = r[SO_IND_COL]
            ind_str = str(int(ind)) if isinstance(ind, float) and ind == int(ind) else str(ind)
            print(f"    {r[PO_SO_ID_COL]:<18} {r['구분']:<5} {ind_str:>8}  "
                  f"{r[SO_SECTOR_COL]:<8} {r['기대 Sector']:<8}  {r['Customer name']}")
        if mismatch_count > 20:
            print(f"    ... 외 {mismatch_count - 20}건")


def create_argument_parser() -> argparse.ArgumentParser:
    """CLI 인자 파서 생성"""
    parser = argparse.ArgumentParser(
        prog='reconcile_ind',
        description='Industry Code 대사 — Orderbook Industry code 채움 + SO Sector 검증',
        epilog='예시:\n'
               '  python reconcile_ind.py P03              # 전체 (채움 + Sector 검증)\n'
               '  python reconcile_ind.py P03 --sector-only  # Sector 검증만',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        'period',
        nargs='?',
        default=None,
        help='대사 월 코드 (예: P03, P04). --sector-only 시 생략 가능',
    )

    parser.add_argument(
        '--sector-only',
        action='store_true',
        help='Sector 검증만 실행 (Industry code 채움 생략, period 불필요)',
    )

    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='상세 로그 출력',
    )

    return parser


def main() -> int:
    """메인 함수"""
    parser = create_argument_parser()
    args = parser.parse_args()

    setup_logging(verbose=args.verbose)
    period = args.period.upper() if args.period else None

    if not args.sector_only and not period:
        print("[오류] period를 입력하세요 (예: python reconcile_ind.py P03)")
        return 1

    # 1. 파일 찾기
    ob_file = find_orderbook_file(period)
    if not ob_file:
        target = f"ind_code_reconciliation/{period}/" if period else "ind_code_reconciliation/*/"
        print(f"[오류] Orderbook 파일을 찾을 수 없습니다 ({target}*orderbook*)")
        return 1

    if not NOAH_SO_PO_DN_FILE.exists():
        print(f"[오류] NOAH_SO_PO_DN.xlsx를 찾을 수 없습니다: {NOAH_SO_PO_DN_FILE}")
        return 1

    print(f"Orderbook: {ob_file.name}")
    print(f"NOAH:      {NOAH_SO_PO_DN_FILE.name}")

    # ── Sector 검증 (항상 실행) ──
    try:
        mismatches, sector_summary = validate_so_sector(ob_file)
    except Exception as e:
        print(f"[오류] Sector 검증 실패: {e}")
        return 1

    sector_file = RECON_DIR / "sector_검증.xlsx"
    write_sector_output(mismatches, sector_summary, sector_file)

    total_null = filled = so_missing = po_missing = dup_review = 0

    # ── Industry code 채움 (--sector-only가 아닐 때) ──
    if not args.sector_only:
        ob_sheet = get_ob_sheet_name(period)
        try:
            ob = pd.read_excel(ob_file, sheet_name=ob_sheet)
        except Exception as e:
            print(f"[오류] Orderbook 로드 실패: {e}")
            return 1

        # Industry code 컬럼 탐색 (typo 대응)
        ind_col = None
        for candidate in ['Indusry code', 'Industry code']:
            if candidate in ob.columns:
                ind_col = candidate
                break
        if ind_col is None:
            print(f"[오류] Industry code 컬럼을 찾을 수 없습니다")
            return 1

        total_null = int(ob[ind_col].isna().sum())

        try:
            mapping, dup_oc_set = build_mapping()
        except Exception as e:
            print(f"[오류] 매핑 데이터 로드 실패: {e}")
            return 1

        print(f"전체 행:   {len(ob)}행, 빈 Industry code: {total_null}행")
        print(f"매핑 건수: {len(mapping)}건 (PO→SO)")

        if total_null > 0:
            result, total_null, filled, so_missing, po_missing, dup_review = fill_industry_code(
                ob, mapping, ind_col, dup_oc_set=dup_oc_set,
            )
        else:
            result = ob.copy()
            result['NOAH Sector'] = None
            result['매핑상태'] = None

        ind_file = RECON_DIR / period / f"ind_code_결과_{period}.xlsx"
        write_ind_code_output(result, ind_file)
        print(f"\n출력1: {ind_file}")

    print(f"{'출력2' if not args.sector_only else '출력'}: {sector_file}")

    # 콘솔 요약
    if not args.sector_only:
        print_ind_summary(total_null, filled, so_missing, po_missing, dup_review)
    print_sector_summary(mismatches, sector_summary)

    return 0


if __name__ == "__main__":
    sys.exit(main())
