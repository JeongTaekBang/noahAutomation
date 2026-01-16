"""
Excel 생성 모듈 (템플릿 기반)
=============================

템플릿 파일을 로드하여 Purchase Order 및 Description 시트를 생성합니다.
사용자는 템플릿 파일에 직접 로고/도장 이미지를 추가할 수 있습니다.
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from po_generator.config import (
    COLORS,
    COLUMN_WIDTHS,
    TOTAL_COLUMNS,
    ITEM_START_ROW,
    SPEC_FIELDS,
    OPTION_FIELDS,
    PO_TEMPLATE_FILE,
)
from po_generator.utils import get_safe_value, get_value, escape_excel_formula
from po_generator.template_engine import (
    load_template,
    generate_po_template,
    clone_row,
    insert_rows_with_template,
    update_sum_formula,
    shift_formula_references,
    copy_cell_style,
)

logger = logging.getLogger(__name__)


# === 스타일 정의 (템플릿 수정 시 사용) ===
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _ensure_template_exists() -> None:
    """템플릿 파일이 없으면 생성"""
    if not PO_TEMPLATE_FILE.exists():
        logger.info("템플릿 파일이 없습니다. 새로 생성합니다...")
        generate_po_template()


def _fill_header_data(
    ws: Worksheet,
    order_data: pd.Series,
    rck_order_no: str,
    today_str: str,
) -> None:
    """헤더 섹션에 데이터 채움 (Row 1-11)

    Args:
        ws: 워크시트
        order_data: 주문 데이터
        rck_order_no: RCK Order No.
        today_str: 오늘 날짜 문자열
    """
    customer_name = get_value(order_data, 'customer_name')
    customer_po = get_value(order_data, 'customer_po')

    # 배송 주소 찾기
    delivery_addr = ''
    for col in order_data.index:
        if '납품' in str(col) or '주소' in str(col):
            val = order_data.get(col, '')
            if pd.notna(val) and str(val) != 'nan':
                delivery_addr = str(val)
                break

    # 데이터 채움 (수식 인젝션 방지 적용)
    ws['A1'] = f"Purchase Order - {escape_excel_formula(rck_order_no)}"
    ws['A5'] = f"Date:  {today_str}"
    ws['C5'] = escape_excel_formula(delivery_addr)
    ws['C7'] = escape_excel_formula(customer_po)
    ws['A10'] = escape_excel_formula(customer_name)


def _fill_item_data(
    ws: Worksheet,
    row_num: int,
    item_idx: int,
    item_data: pd.Series,
    currency: str = 'KRW',
) -> None:
    """아이템 행에 데이터 채움

    Args:
        ws: 워크시트
        row_num: 행 번호
        item_idx: 아이템 인덱스 (0부터 시작)
        item_data: 아이템 데이터
        currency: 통화 코드 (KRW 또는 USD)
    """
    number_format = '₩#,##0' if currency == 'KRW' else '$#,##0.00'

    # 데이터 추출
    model = get_value(item_data, 'model')
    power = get_value(item_data, 'power_supply')
    item_name = get_value(item_data, 'item_name')

    # Description 조합
    desc_parts = []
    if item_name:
        desc_parts.append(item_name)
    elif model:
        desc_parts.append(model)

    if power and isinstance(power, str):
        desc_parts.append(power.replace('-1Ph-', ', ').replace('-3Ph-', ', '))

    if str(get_value(item_data, 'als')).upper() == 'Y':
        desc_parts.append('ALS')

    description = ', '.join([p for p in desc_parts if p])

    # 수량
    try:
        qty = int(float(get_value(item_data, 'item_qty', 1)))
    except (ValueError, TypeError):
        qty = 1

    # 단가
    try:
        ico_unit = float(get_value(item_data, 'ico_unit', 0))
    except (ValueError, TypeError):
        ico_unit = 0

    # 납기일
    requested_date = get_value(item_data, 'delivery_date')
    requested_date_str = ''
    if requested_date and not pd.isna(requested_date):
        try:
            if isinstance(requested_date, datetime):
                requested_date_str = requested_date.strftime("%Y-%m-%d")
            else:
                requested_date_str = str(requested_date)[:10]
        except (ValueError, TypeError):
            requested_date_str = ''

    # 데이터 입력 (수식 인젝션 방지 적용)
    ws[f'A{row_num}'] = item_idx + 1
    ws[f'B{row_num}'] = escape_excel_formula(description)
    ws[f'F{row_num}'] = qty
    ws[f'G{row_num}'] = "EA"
    ws[f'H{row_num}'] = ico_unit
    ws[f'H{row_num}'].number_format = number_format
    ws[f'I{row_num}'] = requested_date_str
    ws[f'J{row_num}'] = f"=H{row_num}*F{row_num}"
    ws[f'J{row_num}'].number_format = number_format


def _fill_footer_data(
    ws: Worksheet,
    order_data: pd.Series,
    footer_start_row: int,
    is_export: bool = False,
) -> int:
    """푸터 섹션에 데이터 채움

    Args:
        ws: 워크시트
        order_data: 주문 데이터
        footer_start_row: 푸터 시작 행 (합계 섹션 다음)
        is_export: 해외 여부

    Returns:
        마지막 행 번호
    """
    remark = get_value(order_data, 'remark')
    incoterms = 'EXW' if is_export else get_value(order_data, 'incoterms')
    currency = 'KRW'

    # 프로젝트 정보
    opportunity = get_value(order_data, 'opportunity')
    sector = get_value(order_data, 'sector')
    industry_code = get_value(order_data, 'industry_code')

    r = footer_start_row
    ws[f'D{r}'] = escape_excel_formula(opportunity)
    r += 1
    ws[f'D{r}'] = escape_excel_formula(sector)
    r += 1
    ws[f'D{r}'] = escape_excel_formula(industry_code)
    r += 1
    ws[f'C{r}'] = f"Note. {escape_excel_formula(remark)}" if remark else "Note."
    r += 1
    ws[f'B{r}'] = currency
    r += 1
    ws[f'B{r}'] = incoterms

    # 마지막 행 (청록색 푸터)은 r + 3
    return r + 3


def _update_print_area(ws: Worksheet, last_row: int) -> None:
    """인쇄 영역 업데이트"""
    ws.print_area = f'A1:J{last_row}'
    ws.row_dimensions[last_row].height = 25


def create_purchase_order(
    ws: Worksheet,
    order_data: pd.Series,
    items_df: Optional[pd.DataFrame] = None,
) -> None:
    """Purchase Order 시트 생성 (템플릿 기반)

    템플릿 파일을 로드하여 데이터를 채웁니다.
    템플릿이 없으면 자동 생성합니다.

    Args:
        ws: 워크시트 (템플릿에서 복사된 시트)
        order_data: 첫 번째 아이템 데이터 (공통 정보 추출용)
        items_df: 다중 아이템인 경우 DataFrame, 단일이면 None
    """
    logger.info("Purchase Order 시트 생성 중 (템플릿 기반)...")

    # 아이템 목록 준비
    if items_df is not None:
        items_list = [row for _, row in items_df.iterrows()]
    else:
        items_list = [order_data]

    num_items = len(items_list)

    # 공통 데이터
    rck_order_no = get_value(order_data, 'order_no')
    today_str = datetime.now().strftime("%d/%b/%Y").upper()
    currency = 'KRW'

    # 해외(수출) 건 여부 확인
    sheet_type = get_value(order_data, 'sheet_type', '')
    is_export = sheet_type == '해외'

    # 1. 헤더 데이터 채움 (Row 1-11)
    _fill_header_data(ws, order_data, rck_order_no, today_str)

    # 2. 아이템 행 처리
    template_row = ITEM_START_ROW  # Row 13
    rows_to_insert = num_items - 1

    if rows_to_insert > 0:
        # 추가 행 삽입 (Row 14부터)
        ws.insert_rows(template_row + 1, rows_to_insert)

        # 삽입된 행에 템플릿 스타일 복제
        for i in range(rows_to_insert):
            target_row = template_row + 1 + i
            clone_row(ws, template_row, target_row, TOTAL_COLUMNS)

    # 아이템 데이터 채움
    for item_idx, item_data in enumerate(items_list):
        row_num = template_row + item_idx
        _fill_item_data(ws, row_num, item_idx, item_data, currency)

    item_last_row = template_row + num_items - 1

    # 3. 합계 섹션 업데이트 (동적 위치)
    # 템플릿에서 합계 섹션은 Row 14-16이었음
    # 아이템 삽입 후 위치가 변경됨
    totals_start_row = item_last_row + 1
    row_total_net = totals_start_row
    row_vat = totals_start_row + 1
    row_order_total = totals_start_row + 2

    # SUM 공식 범위 업데이트
    ws[f'J{row_total_net}'] = f"=SUM(J{ITEM_START_ROW}:J{item_last_row})"

    # VAT 처리 (해외는 0)
    if is_export:
        ws[f'J{row_vat}'] = 0
    else:
        ws[f'J{row_vat}'] = f"=J{row_total_net}*0.1"

    # Order Total 공식 업데이트
    ws[f'J{row_order_total}'] = f"=SUM(J{row_total_net}:J{row_vat})"

    # 4. 푸터 데이터 채움
    footer_start_row = row_order_total + 1
    last_row = _fill_footer_data(ws, order_data, footer_start_row, is_export)

    # 5. 인쇄 영역 업데이트
    _update_print_area(ws, last_row)

    logger.info(f"Purchase Order 시트 생성 완료 (아이템 {num_items}개)")


def create_description_sheet(
    ws: Worksheet,
    order_data: pd.Series,
    items_df: Optional[pd.DataFrame] = None,
) -> None:
    """Description 시트 생성 (템플릿 기반)

    Args:
        ws: 워크시트 (템플릿에서 복사된 시트)
        order_data: 첫 번째 아이템 데이터
        items_df: 다중 아이템인 경우 DataFrame, 단일이면 None
    """
    logger.info("Description 시트 생성 중 (템플릿 기반)...")

    # 아이템 목록 준비
    if items_df is not None:
        items_list = [row for _, row in items_df.iterrows()]
    else:
        items_list = [order_data]

    num_items = len(items_list)

    # 추가 아이템 열 생성 (B열은 이미 템플릿에 있음)
    if num_items > 1:
        # C열부터 추가
        for idx in range(1, num_items):
            col_num = 2 + idx  # C, D, E...
            col_letter = get_column_letter(col_num)

            # Row 1: Line No
            ws.cell(row=1, column=col_num, value=idx + 1)
            ws.cell(row=1, column=col_num).border = THIN_BORDER
            ws.cell(row=1, column=col_num).alignment = Alignment(horizontal='center')

            # Row 2: Qty
            try:
                qty = int(float(get_value(items_list[idx], 'item_qty', 1)))
            except (ValueError, TypeError):
                qty = 1
            ws.cell(row=2, column=col_num, value=qty)
            ws.cell(row=2, column=col_num).border = THIN_BORDER
            ws.cell(row=2, column=col_num).alignment = Alignment(horizontal='center')

            # 열 너비
            ws.column_dimensions[col_letter].width = 15

    # 첫 번째 아이템 데이터 (B열)
    try:
        qty_first = int(float(get_value(items_list[0], 'item_qty', 1)))
    except (ValueError, TypeError):
        qty_first = 1
    ws['B2'] = qty_first

    # SPEC_FIELDS 데이터 채움 (수식 인젝션 방지 적용)
    row_idx = 3
    for field in SPEC_FIELDS:
        for idx, item_data in enumerate(items_list):
            col_num = 2 + idx
            value = get_safe_value(item_data, field)
            escaped_value = escape_excel_formula(value) if value else None
            ws.cell(row=row_idx, column=col_num, value=escaped_value)
            ws.cell(row=row_idx, column=col_num).border = THIN_BORDER
        row_idx += 1

    # OPTION_FIELDS 데이터 채움 (수식 인젝션 방지 적용)
    for field in OPTION_FIELDS:
        for idx, item_data in enumerate(items_list):
            col_num = 2 + idx
            value = get_safe_value(item_data, field)
            escaped_value = escape_excel_formula(value) if value else None
            ws.cell(row=row_idx, column=col_num, value=escaped_value)
            ws.cell(row=row_idx, column=col_num).border = THIN_BORDER
        row_idx += 1

    logger.info("Description 시트 생성 완료")


def create_po_workbook(
    order_data: pd.Series,
    items_df: Optional[pd.DataFrame] = None,
) -> Workbook:
    """템플릿 기반으로 PO Workbook 생성

    Args:
        order_data: 주문 데이터
        items_df: 다중 아이템 DataFrame (선택)

    Returns:
        생성된 Workbook
    """
    # 템플릿 확인 및 로드
    _ensure_template_exists()
    wb = load_template()

    # Purchase Order 시트
    ws_po = wb['Purchase Order']
    create_purchase_order(ws_po, order_data, items_df)

    # Description 시트
    ws_desc = wb['Description']
    create_description_sheet(ws_desc, order_data, items_df)

    return wb
