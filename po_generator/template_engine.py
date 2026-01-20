"""
템플릿 엔진 모듈 (Deprecated)
=============================

이 모듈은 openpyxl 기반 템플릿 처리를 위해 사용되었습니다.
현재는 xlwings 기반 excel_generator.py로 대체되었습니다.

generate_po_template() 함수만 유지하여 템플릿 생성 기능을 제공합니다.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from po_generator.config import (
    PO_TEMPLATE_FILE,
    TEMPLATE_DIR,
    COLORS,
    COLUMN_WIDTHS,
    TOTAL_COLUMNS,
    SPEC_FIELDS,
    OPTION_FIELDS,
)

logger = logging.getLogger(__name__)


def ensure_template_dir() -> None:
    """템플릿 디렉토리가 없으면 생성"""
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)


def generate_po_template() -> Path:
    """Purchase Order 템플릿 파일 생성

    현재 코드 기반 레이아웃으로 템플릿을 생성합니다.
    이미지는 사용자가 직접 추가할 수 있습니다.

    Returns:
        생성된 템플릿 파일 경로
    """
    ensure_template_dir()

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    # 스타일 정의
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    red_fill = PatternFill(start_color=COLORS.RED, end_color=COLORS.RED, fill_type="solid")
    gray_fill = PatternFill(start_color=COLORS.GRAY, end_color=COLORS.GRAY, fill_type="solid")
    teal_fill = PatternFill(start_color=COLORS.TEAL, end_color=COLORS.TEAL, fill_type="solid")
    green_fill = PatternFill(start_color=COLORS.GREEN, end_color=COLORS.GREEN, fill_type="solid")
    red_bright_fill = PatternFill(
        start_color=COLORS.RED_BRIGHT, end_color=COLORS.RED_BRIGHT, fill_type="solid"
    )

    white_font = Font(color="FFFFFF")
    white_bold_font = Font(bold=True, color="FFFFFF")

    # === Row 1: 타이틀 (빨간 배경) ===
    ws['A1'] = "Purchase Order - "  # RCK Order no. 가 붙음
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].alignment = Alignment(vertical='center')
    for col in range(1, TOTAL_COLUMNS + 1):
        ws.cell(row=1, column=col).fill = red_fill

    # === Row 2-3: Vendor Info + Rotork 로고 ===
    ws['A2'] = "Vendor Name:  NOAH Actuation"
    ws['A2'].font = Font(size=10)
    ws['J2'] = "rotork"
    ws['J2'].font = Font(bold=True, size=14, color=COLORS.RED, italic=True)
    ws['J2'].alignment = Alignment(horizontal='right', vertical='center')

    ws['A3'] = "Vendor Number: -"
    ws['A3'].font = Font(size=10)
    ws['J3'] = "Rotork Korea"
    ws['J3'].font = Font(size=10)
    ws['J3'].alignment = Alignment(horizontal='right', vertical='top')

    # === Row 4-5: Vendor Reference + Delivery Address ===
    ws['A4'] = "Vendor Reference: -"
    ws['A4'].font = Font(size=10)
    ws['C4'] = "Delivery Address:"
    ws['C4'].font = Font(size=9, bold=True, italic=True)

    ws['A5'] = "Date:  "  # 날짜가 붙음
    ws['A5'].font = Font(size=10)
    ws['C5'] = ""  # Delivery Address 값
    ws['C5'].font = Font(size=9, italic=True)
    ws['C5'].alignment = Alignment(wrap_text=True)

    # === Row 6-11: Supplier/Customer Info ===
    ws['A6'] = "Supplier address:"
    ws['A6'].font = Font(size=9, bold=True)
    ws['C6'] = "Customer PO:"
    ws['C6'].font = Font(size=9, bold=True)

    ws['A7'] = "NOAH Actuation"
    ws['A7'].font = Font(size=9)
    ws['C7'] = ""  # Customer PO 값
    ws['C7'].font = Font(size=9)

    ws['A8'] = "11 Jeongseojin-9ro, Seo-gu, Incheon, Korea(22850)"
    ws['A8'].font = Font(size=9)

    ws['A9'] = "Final Customer"
    ws['A9'].font = Font(size=9, bold=True)

    ws['A10'] = ""  # Customer name 값
    ws['A10'].font = Font(size=9)

    ws['A11'] = "Order Details"
    ws['A11'].font = Font(size=9, bold=True)

    # === Row 12: 아이템 헤더 (회색 배경) ===
    ws.merge_cells('B12:E12')
    headers = [
        ('A12', 'Item\nNumber'),
        ('B12', 'Description'),
        ('F12', 'Qty'),
        ('G12', 'Unit'),
        ('H12', 'Unit price'),
        ('I12', 'Delivery\nRequired'),
        ('J12', 'Amount'),
    ]
    header_font = Font(bold=True, size=10, color="FFFFFF")
    for cell_addr, text in headers:
        ws[cell_addr] = text
        ws[cell_addr].font = header_font
        ws[cell_addr].fill = gray_fill
        ws[cell_addr].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center'
        )
        ws[cell_addr].border = thin_border

    for col in range(1, TOTAL_COLUMNS + 1):
        ws.cell(row=12, column=col).fill = gray_fill
        ws.cell(row=12, column=col).border = thin_border

    # === Row 13: 아이템 템플릿 행 (복제될 행) ===
    ws.merge_cells('B13:E13')
    ws['A13'] = 1
    ws['A13'].alignment = Alignment(horizontal='center')
    ws['B13'] = ""  # Description
    ws['F13'] = 1   # Qty
    ws['F13'].alignment = Alignment(horizontal='center')
    ws['G13'] = "EA"
    ws['G13'].alignment = Alignment(horizontal='center')
    ws['H13'] = 0   # Unit price
    ws['H13'].number_format = '₩#,##0'
    ws['H13'].alignment = Alignment(horizontal='right')
    ws['I13'] = ""  # Delivery date
    ws['I13'].alignment = Alignment(horizontal='center')
    ws['J13'] = "=H13*F13"  # Amount 공식
    ws['J13'].number_format = '₩#,##0'
    ws['J13'].alignment = Alignment(horizontal='right')

    for col in range(1, TOTAL_COLUMNS + 1):
        ws.cell(row=13, column=col).border = thin_border

    # === Row 14-16: 합계 섹션 ===
    title_font = Font(bold=True, size=11)

    ws['I14'] = "Total net amount"
    ws['I14'].font = title_font
    ws['I14'].alignment = Alignment(horizontal='right')
    ws['J14'] = "=SUM(J13:J13)"  # 동적으로 조정됨
    ws['J14'].number_format = '₩#,##0'
    ws['J14'].alignment = Alignment(horizontal='right')
    ws['J14'].border = thin_border

    ws['I15'] = "VAT"
    ws['I15'].alignment = Alignment(horizontal='right')
    ws['J15'] = "=J14*0.1"  # 해외는 0으로 변경됨
    ws['J15'].number_format = '₩#,##0'
    ws['J15'].alignment = Alignment(horizontal='right')
    ws['J15'].border = thin_border

    ws['I16'] = "Order Total"
    ws['I16'].font = title_font
    ws['I16'].alignment = Alignment(horizontal='right')
    ws['J16'] = "=SUM(J14:J15)"
    ws['J16'].number_format = '₩#,##0'
    ws['J16'].font = Font(bold=True)
    ws['J16'].alignment = Alignment(horizontal='right')
    ws['J16'].border = thin_border

    # === Row 17-25: 푸터 섹션 ===
    r = 17
    ws[f'A{r}'] = "D365CEProject:"
    ws[f'C{r}'] = "Opportunity:"
    ws[f'D{r}'] = ""  # Opportunity 값

    r += 1
    ws[f'A{r}'] = "Project name:"
    ws[f'C{r}'] = "Sector:"
    ws[f'D{r}'] = ""  # Sector 값

    r += 1
    ws[f'C{r}'] = "Industry code:"
    ws[f'D{r}'] = ""  # Industry code 값

    r += 1
    ws[f'A{r}'] = "Additional information"
    ws[f'C{r}'] = "Note."  # Remark가 붙음
    ws[f'C{r}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws[f'H{r}'] = "On behalf of Rotork"

    r += 1
    ws[f'A{r}'] = "Order Currency:"
    ws[f'B{r}'] = "KRW"  # Currency 값
    ws[f'H{r}'] = "Contact:"

    r += 1
    ws[f'A{r}'] = "Delivery Terms:"
    ws[f'B{r}'] = ""  # Incoterms 값

    r += 1
    ws[f'A{r}'] = "Delivery mode:"
    ws[f'H{r}'] = "Email:"

    r += 1
    ws[f'A{r}'] = "Terms of payment:"
    ws[f'H{r}'] = "Tel:"

    r += 1  # Row 25: 청록색 푸터
    ws[f'A{r}'] = "Keeping the World flowing for Future Generations"
    ws[f'A{r}'].font = Font(size=12, color="FFFFFF", italic=True)
    ws[f'A{r}'].alignment = Alignment(vertical='center')
    ws[f'J{r}'] = "1 of 1"
    ws[f'J{r}'].font = Font(size=12, color="FFFFFF")
    ws[f'J{r}'].alignment = Alignment(horizontal='right', vertical='center')

    for col in range(1, TOTAL_COLUMNS + 1):
        ws.cell(row=r, column=col).fill = teal_fill

    # === 레이아웃 설정 ===
    for col, width in COLUMN_WIDTHS.as_dict().items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[12].height = 30
    ws.row_dimensions[r].height = 25

    ws.print_area = f'A1:J{r}'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = 'portrait'
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # === Description 시트 템플릿 ===
    ws_desc = wb.create_sheet("Description")

    # Row 1: Line No
    ws_desc['A1'] = "Line No"
    ws_desc['A1'].font = white_bold_font
    ws_desc['A1'].fill = green_fill
    ws_desc['A1'].border = thin_border
    ws_desc['A1'].alignment = Alignment(horizontal='center')

    ws_desc['B1'] = 1  # 첫 번째 아이템
    ws_desc['B1'].border = thin_border
    ws_desc['B1'].alignment = Alignment(horizontal='center')

    # Row 2: Qty
    ws_desc['A2'] = "Qty"
    ws_desc['A2'].font = white_bold_font
    ws_desc['A2'].fill = green_fill
    ws_desc['A2'].border = thin_border
    ws_desc['A2'].alignment = Alignment(horizontal='center')

    ws_desc['B2'] = 1
    ws_desc['B2'].border = thin_border
    ws_desc['B2'].alignment = Alignment(horizontal='center')

    # Row 3+: SPEC_FIELDS (초록 배경)
    row_idx = 3
    for field in SPEC_FIELDS:
        ws_desc.cell(row=row_idx, column=1, value=field)
        ws_desc.cell(row=row_idx, column=1).font = white_bold_font
        ws_desc.cell(row=row_idx, column=1).fill = green_fill
        ws_desc.cell(row=row_idx, column=1).border = thin_border
        ws_desc.cell(row=row_idx, column=2).border = thin_border
        row_idx += 1

    # OPTION_FIELDS (빨간 배경)
    for field in OPTION_FIELDS:
        ws_desc.cell(row=row_idx, column=1, value=field)
        ws_desc.cell(row=row_idx, column=1).font = white_bold_font
        ws_desc.cell(row=row_idx, column=1).fill = red_bright_fill
        ws_desc.cell(row=row_idx, column=1).border = thin_border
        ws_desc.cell(row=row_idx, column=2).border = thin_border
        row_idx += 1

    # 열 너비
    ws_desc.column_dimensions['A'].width = 25
    ws_desc.column_dimensions['B'].width = 15

    # 저장
    wb.save(PO_TEMPLATE_FILE)
    logger.info(f"PO 템플릿 생성 완료: {PO_TEMPLATE_FILE}")

    return PO_TEMPLATE_FILE


# === 아래 함수들은 Deprecated - 하위 호환성을 위해 유지 ===

def load_template(template_path: Path | None = None):
    """템플릿 파일 로드 (Deprecated - xlwings로 대체됨)"""
    from openpyxl import load_workbook
    if template_path is None:
        template_path = PO_TEMPLATE_FILE
    if not template_path.exists():
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
    return load_workbook(template_path, data_only=False)


def clone_row(ws, source_row, target_row, max_col=10):
    """행 복제 (Deprecated - xlwings로 대체됨)"""
    pass


def insert_rows_with_template(ws, template_row, count, max_col=10):
    """템플릿 행 복제하여 삽입 (Deprecated - xlwings로 대체됨)"""
    return template_row + count - 1 if count > 1 else template_row


def update_sum_formula(ws, cell_address, item_start_row, item_end_row, column='J'):
    """SUM 공식 범위 업데이트 (Deprecated - xlwings로 대체됨)"""
    pass


def shift_formula_references(formula, shift_by, reference_start_row):
    """공식 행 참조 이동 (Deprecated - xlwings로 대체됨)"""
    return formula


def copy_cell_style(source, target):
    """셀 스타일 복사 (Deprecated - xlwings로 대체됨)"""
    pass
