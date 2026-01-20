"""
이력 관리 모듈
==============

발주서 생성 이력을 관리합니다.
- 중복 발주 체크
- 이력 저장 (발주서 파일에서 데이터 추출 → DB 형식으로 저장)
- 발주서 기준 스냅샷 보존
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, TypedDict
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from po_generator.config import (
    HISTORY_DIR,
    ITEM_START_ROW_FALLBACK,
)

logger = logging.getLogger(__name__)


# === 이력 추출용 로컬 상수 (config.py에서 이동됨) ===
# 헤더 영역 고정 위치 (Purchase Order 시트)
CELL_TITLE: str = "A1"
CELL_DATE: str = "A5"
CELL_CUSTOMER_NAME: str = "A10"

# 이력 아이템 검색 제한 (기본값)
HISTORY_MAX_SEARCH_ROWS: int = 100


def _find_item_start_row(
    ws,
    search_labels: tuple[str, ...] = ('No.', 'Item Number', 'Item\nNumber', '품명', 'Item'),
    max_search_rows: int = 30,
    fallback_row: int = ITEM_START_ROW_FALLBACK,
) -> int:
    """템플릿에서 아이템 시작 행을 동적으로 찾기 (openpyxl 버전)

    헤더 레이블을 찾아서 그 다음 행이 아이템 시작 위치입니다.

    Args:
        ws: openpyxl Worksheet 객체
        search_labels: 검색할 헤더 레이블
        max_search_rows: 최대 검색 행 수
        fallback_row: 헤더를 찾지 못했을 때 기본값

    Returns:
        아이템 시작 행 번호
    """
    for row in range(1, max_search_rows + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell_value = ws[f'{col}{row}'].value
            if cell_value and any(
                label in str(cell_value) for label in search_labels
            ):
                logger.debug(f"헤더 발견: Row {row}, 값='{cell_value}' -> 아이템 시작 Row {row + 1}")
                return row + 1

    logger.debug(f"헤더를 찾지 못함 -> 기본값 Row {fallback_row} 사용")
    return fallback_row


class DuplicateInfo(TypedDict):
    """중복 발주 정보"""
    생성일시: str
    생성파일: str


def sanitize_filename(name: str) -> str:
    r"""파일명에 사용할 수 없는 문자 제거

    Windows 파일명 금지 문자(\ / : * ? " < > |)를 제거하고
    연속 공백/언더스코어를 정리합니다.

    Args:
        name: 정규화할 파일명 또는 문자열

    Returns:
        안전한 파일명 문자열
    """
    # Windows 파일명 금지 문자: \ / : * ? " < > |
    sanitized = re.sub(r'[\\/:*?"<>|]', '_', str(name))
    # 연속 공백/언더스코어 정리
    sanitized = re.sub(r'[_\s]+', '_', sanitized)
    return sanitized.strip('_')


# 하위 호환성을 위한 별칭
_sanitize_filename = sanitize_filename


def _get_current_month_dir() -> Path:
    """현재 연/월 폴더 경로 반환: po_history/YYYY/M월/"""
    now = datetime.now()
    return HISTORY_DIR / str(now.year) / f"{now.month}월"


def _ensure_history_dir() -> Path:
    """이력 폴더 생성 (없으면) - 월별 폴더"""
    month_dir = _get_current_month_dir()
    month_dir.mkdir(parents=True, exist_ok=True)
    return month_dir


def _get_history_filename(order_no: str, customer_name: str) -> str:
    """이력 파일명 생성: YYYYMMDD_주문번호_고객명.xlsx

    Args:
        order_no: RCK Order No.
        customer_name: 고객명

    Returns:
        생성된 파일명 (예: "20260116_ND-0001_고객명.xlsx")
    """
    date_str = datetime.now().strftime("%Y%m%d")
    safe_order = _sanitize_filename(order_no)
    safe_customer = _sanitize_filename(customer_name)
    return f"{date_str}_{safe_order}_{safe_customer}.xlsx"


def _extract_data_from_po_file(po_file: Path) -> dict[str, Any]:
    """발주서 파일에서 데이터 추출 (DB 형식) - openpyxl 버전

    Purchase Order 시트와 Description 시트에서 데이터를 추출하여
    한 행의 딕셔너리로 반환합니다.
    아이템 수에 따라 동적으로 셀 위치를 계산합니다.

    Args:
        po_file: 발주서 파일 경로

    Returns:
        추출된 데이터 딕셔너리
    """
    record: dict[str, Any] = {}

    try:
        wb = load_workbook(po_file, data_only=True)

        # === Purchase Order 시트에서 추출 ===
        if "Purchase Order" in wb.sheetnames:
            ws_po = wb["Purchase Order"]

            # 아이템 시작 행 동적 탐지
            item_start_row = _find_item_start_row(ws_po, fallback_row=ITEM_START_ROW_FALLBACK)

            # 고정 위치 셀 (헤더 영역)
            # A1: "Purchase Order - ND-0001" 형식에서 주문번호 추출
            title = ws_po['A1'].value or ''
            if ' - ' in str(title):
                record['RCK Order no.'] = title.split(' - ', 1)[1]

            # A10: 고객명
            record['Customer name'] = ws_po['A10'].value

            # A5: 날짜 "Date:  01/JAN/2026"
            date_cell = ws_po['A5'].value or ''
            if 'Date:' in str(date_cell):
                record['PO Date'] = str(date_cell).replace('Date:', '').strip()

            # 아이템 정보 동적 추출 (동적 탐지된 행부터 빈 셀까지)
            items_data = []
            row = item_start_row
            searched = 0

            while searched < HISTORY_MAX_SEARCH_ROWS:
                desc = ws_po[f'B{row}'].value
                if not desc:  # 빈 셀이면 아이템 끝
                    break

                qty = ws_po[f'F{row}'].value
                unit_price = ws_po[f'H{row}'].value
                delivery_date = ws_po[f'I{row}'].value

                items_data.append({
                    'Description': desc,
                    'Item qty': qty,
                    'ICO Unit': unit_price,
                    'Requested delivery date': delivery_date,
                })
                row += 1
                searched += 1

            # 아이템 마지막 행 (동적 계산)
            item_last_row = item_start_row + len(items_data) - 1 if items_data else item_start_row

            # 첫 번째 아이템 정보를 메인 레코드에
            if items_data:
                record['Item count'] = len(items_data)
                record['Description'] = items_data[0].get('Description')
                record['Item qty'] = items_data[0].get('Item qty')
                record['ICO Unit'] = items_data[0].get('ICO Unit')
                record['Requested delivery date'] = items_data[0].get('Requested delivery date')

            # 합계 행 동적 계산 (아이템 마지막 행 + 1)
            totals_start_row = item_last_row + 1
            row_total_net = totals_start_row
            row_order_total = totals_start_row + 2

            # J{row}: Total net amount (동적 위치)
            record['Total net amount'] = ws_po[f'J{row_total_net}'].value

            # J{row+2}: Order Total (VAT 포함, 동적 위치)
            record['Order Total'] = ws_po[f'J{row_order_total}'].value

            # 푸터 행 동적 계산 (합계 행 + 4에서 시작)
            footer_start_row = row_order_total + 1
            currency_row = footer_start_row + 4
            incoterms_row = footer_start_row + 5

            # B{row}: Currency (동적 위치)
            record['Currency'] = ws_po[f'B{currency_row}'].value

            # B{row}: Incoterms (동적 위치)
            record['Incoterms'] = ws_po[f'B{incoterms_row}'].value

        # === Description 시트에서 추출 ===
        if "Description" in wb.sheetnames:
            ws_desc = wb["Description"]

            # Description 시트는 세로 형식 (A열: 필드명, B열: 값)
            for row in range(2, min(ws_desc.max_row + 1, 100)):  # 최대 100행까지
                field_name = ws_desc[f'A{row}'].value
                field_value = ws_desc[f'B{row}'].value

                if field_name and field_name != "Line No":
                    record[field_name] = field_value

        wb.close()

    except BadZipFile as e:
        logger.error(f"발주서 파일 손상: {e}")
    except InvalidFileException as e:
        logger.error(f"발주서 파일 형식 오류: {e}")
    except PermissionError as e:
        logger.error(f"발주서 파일 접근 권한 없음: {e}")
    except KeyError as e:
        logger.error(f"발주서 시트/셀 없음: {e}")
    except ValueError as e:
        logger.error(f"발주서 데이터 형식 오류: {e}")
    except Exception as e:
        logger.error(f"발주서 파일 읽기 오류: {e}")

    return record


def check_duplicate_order(order_no: str, check_all_months: bool = True) -> DuplicateInfo | None:
    """중복 발주 체크

    Args:
        order_no: RCK Order No.
        check_all_months: True면 전체 이력 검색, False면 현재 월만

    Returns:
        중복인 경우 이전 발주 정보, 아니면 None
    """
    safe_order_no = _sanitize_filename(order_no)

    # 검색할 폴더 목록 결정
    if check_all_months and HISTORY_DIR.exists():
        # 전체 이력 폴더에서 검색 (po_history/**/*)
        search_dirs = list(HISTORY_DIR.rglob("*월"))
    else:
        # 현재 월만 검색
        month_dir = _get_current_month_dir()
        search_dirs = [month_dir] if month_dir.exists() else []

    for month_dir in search_dirs:
        if not month_dir.is_dir():
            continue

        for history_file in month_dir.glob("*.xlsx"):
            # 파일명에서 주문번호 추출 (YYYYMMDD_주문번호_고객명.xlsx)
            filename = history_file.stem
            parts = filename.split('_', 2)  # 최대 3개로 분리
            if len(parts) >= 2:
                file_order_no = parts[1]
                if file_order_no == safe_order_no:
                    logger.warning(f"중복 발주 감지: {order_no}")
                    # 파일에서 생성일시 읽기
                    try:
                        df = pd.read_excel(history_file)
                        if not df.empty and '생성일시' in df.columns:
                            return DuplicateInfo(
                                생성일시=str(df.iloc[0]['생성일시']),
                                생성파일=str(history_file)
                            )
                    except (InvalidFileException, BadZipFile, PermissionError, ValueError) as e:
                        logger.debug(f"이력 파일 읽기 실패 (파일 시간으로 대체): {e}")
                    # 파일 수정 시간으로 대체
                    mtime = datetime.fromtimestamp(history_file.stat().st_mtime)
                    return DuplicateInfo(
                        생성일시=mtime.strftime("%Y-%m-%d %H:%M:%S"),
                        생성파일=str(history_file)
                    )

    return None


def save_to_history(output_file: Path, order_no: str, customer_name: str) -> bool:
    """발주 이력 저장 (발주서에서 데이터 추출 → DB 형식 저장)

    생성된 발주서 파일에서 데이터를 추출하여 DB 형식(한 행)으로 저장합니다.
    파일명: YYYYMMDD_주문번호_고객명.xlsx

    Args:
        output_file: 생성된 발주서 파일 경로
        order_no: RCK Order No.
        customer_name: 고객명

    Returns:
        저장 성공 여부
    """
    if not output_file.exists():
        logger.error(f"발주서 파일이 없습니다: {output_file}")
        return False

    month_dir = _ensure_history_dir()

    # 발주서에서 데이터 추출
    record = _extract_data_from_po_file(output_file)

    # 메타 정보 추가 (맨 앞에 배치)
    meta_info = {
        '생성일시': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'RCK Order no.': order_no,
        'Customer name': customer_name,
        '원본파일': output_file.name,
    }

    # 메타 정보를 앞에, 추출된 데이터를 뒤에
    full_record = {**meta_info, **record}

    # 파일명 생성 (월별 폴더에 저장)
    filename = _get_history_filename(order_no, customer_name)
    history_file = month_dir / filename

    # 동일 날짜에 같은 주문이 있으면 번호 추가
    counter = 1
    while history_file.exists():
        base_filename = _get_history_filename(order_no, customer_name)
        name_without_ext = base_filename.rsplit('.', 1)[0]
        filename = f"{name_without_ext}_{counter}.xlsx"
        history_file = month_dir / filename
        counter += 1

    try:
        # DB 형식으로 저장 (한 행)
        df = pd.DataFrame([full_record])
        df.to_excel(history_file, index=False)
        logger.info(f"이력 저장 완료: {history_file.name}")
        return True

    except (IOError, OSError, PermissionError) as e:
        logger.error(f"이력 저장 실패 (파일 접근 오류): {e}")
        return False
    except ValueError as e:
        logger.error(f"이력 저장 실패 (데이터 형식 오류): {e}")
        return False


def get_history_count() -> int:
    """이력 건수 조회 (현재 월 폴더만)

    Returns:
        이력 건수
    """
    month_dir = _get_current_month_dir()

    if month_dir.exists():
        return len(list(month_dir.glob("*.xlsx")))

    return 0


def get_all_history() -> pd.DataFrame:
    """현재 월 이력 조회 (월별 폴더 내 모든 파일 합치기)

    현재 월 폴더의 이력 파일(DB 형식)을 읽어서 하나의 DataFrame으로 합칩니다.

    Returns:
        현재 월 이력 DataFrame
    """
    all_records = []
    month_dir = _get_current_month_dir()

    if month_dir.exists():
        for history_file in sorted(month_dir.glob("*.xlsx")):
            try:
                df = pd.read_excel(history_file)
                all_records.append(df)
            except (InvalidFileException, BadZipFile, PermissionError) as e:
                logger.warning(f"이력 파일 읽기 실패 (파일 손상/권한): {history_file.name} - {e}")
            except ValueError as e:
                logger.warning(f"이력 파일 읽기 실패 (데이터 형식): {history_file.name} - {e}")

    if all_records:
        return pd.concat(all_records, ignore_index=True)
    return pd.DataFrame()


def clear_history() -> bool:
    """이력 초기화 - 현재 월 폴더만 (테스트용)

    Returns:
        성공 여부
    """
    try:
        month_dir = _get_current_month_dir()

        if month_dir.exists():
            for f in month_dir.glob("*.xlsx"):
                f.unlink()
            logger.info(f"이력 폴더 내 파일 삭제됨: {month_dir}")

        return True
    except (IOError, OSError, PermissionError) as e:
        logger.error(f"이력 삭제 실패: {e}")
        return False


def get_current_month_info() -> tuple[str, Path]:
    """현재 월 정보 반환 (외부용)

    Returns:
        (월 표시 문자열, 월 폴더 경로)
    """
    now = datetime.now()
    month_str = f"{now.year}년 {now.month}월"
    return month_str, _get_current_month_dir()
